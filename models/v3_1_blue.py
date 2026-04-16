import os
import sys
import glob
import shutil
import logging
import re
import warnings
warnings.filterwarnings("ignore")
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.model_selection import train_test_split, RandomizedSearchCV, StratifiedKFold
from sklearn.base import clone
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from sklearn.impute import SimpleImputer
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics import roc_auc_score, accuracy_score, f1_score, precision_score, recall_score
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import ExtraTreesClassifier
import torch
from transformers import BertTokenizer, BertModel
from tqdm import tqdm
import matplotlib
matplotlib.use('Agg')


# Optional libs
try:
    import lightgbm as lgb
except Exception:
    lgb = None

try:
    import shap
except Exception:
    shap = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
except Exception:
    load_workbook = None
    Font = PatternFill = Alignment = Border = Side = ColorScaleRule = get_column_letter = None

# -----------------------
# 核心配置（统一输出到当前代码目录）
# -----------------------
# 获取当前代码文件所在目录（关键：所有输出都存这里）
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
# 数据路径（改为当前目录相对路径，方便移植）
DATA_PATH = os.path.join(CURRENT_DIR, "WA_Fn-UseC_-HR-Employee-Attrition.csv")
POLICY_PATH = os.path.join(CURRENT_DIR, "人才政策信息表(1).xlsx")
# 其他配置
RANDOM_STATE = 42
TEST_SIZE = 0.2
TIME_DECAY_HALF_LIFE_DAYS = 180
DROP_COLS = ["EmployeeNumber", "Over18", "StandardHours", "DailyRate", "HourlyRate"]

# 日志配置
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s", datefmt="%H:%M:%S")
# Joblib临时目录（放入当前目录，避免权限问题）
os.makedirs(os.path.join(CURRENT_DIR, "temp_joblib"), exist_ok=True)
os.environ["JOBLIB_TEMP_FOLDER"] = os.path.join(CURRENT_DIR, "temp_joblib")

# 可视化全局设置（解决中文乱码、图表美观）
plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False
plt.style.use('seaborn-v0_8-whitegrid')


# -----------------------
# 运行时配置（支持脚本调用 / Web调用）
# -----------------------
def configure_runtime_paths(current_dir=None, data_path=None, policy_path=None):
    """更新运行时路径配置，便于外部封装调用。"""
    global CURRENT_DIR, DATA_PATH, POLICY_PATH

    if current_dir:
        CURRENT_DIR = os.path.abspath(current_dir)
        os.makedirs(CURRENT_DIR, exist_ok=True)
    if data_path:
        DATA_PATH = os.path.abspath(data_path)
    if policy_path:
        POLICY_PATH = os.path.abspath(policy_path)

    temp_dir = os.path.join(CURRENT_DIR, "temp_joblib")
    os.makedirs(temp_dir, exist_ok=True)
    os.environ["JOBLIB_TEMP_FOLDER"] = temp_dir


def safe_float(value):
    """尽量将值转为float，失败返回None。"""
    try:
        return float(value)
    except Exception:
        return None


def style_excel_workbook(file_path, percent_cols_map=None, heatmap_cols_map=None):
    """为Excel文件添加更易读的样式（表头、筛选、冻结、列宽、百分比格式）。"""
    if load_workbook is None:
        logging.warning("openpyxl样式模块不可用，跳过Excel美化：%s", file_path)
        return

    percent_cols_map = percent_cols_map or {}
    heatmap_cols_map = heatmap_cols_map or {}

    try:
        wb = load_workbook(file_path)
    except Exception as exc:
        logging.warning("Excel加载失败，跳过美化：%s | %s", file_path, exc)
        return

    header_fill = PatternFill("solid", fgColor="1F6FEB")
    header_font = Font(color="FFFFFF", bold=True)
    even_fill = PatternFill("solid", fgColor="F7FBFF")
    thin_side = Side(style="thin", color="DCE6F5")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for ws in wb.worksheets:
        max_row, max_col = ws.max_row, ws.max_column
        if max_row < 1 or max_col < 1:
            continue

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # 表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # 数据行样式
        for row_idx in range(2, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if row_idx % 2 == 0:
                    cell.fill = even_fill

        header_to_index = {}
        for col_idx in range(1, max_col + 1):
            header_name = clean_text(ws.cell(row=1, column=col_idx).value)
            if header_name:
                header_to_index[header_name] = col_idx

        # 百分比列格式
        for col_name in percent_cols_map.get(ws.title, []):
            col_idx = header_to_index.get(col_name)
            if not col_idx:
                continue
            for row_idx in range(2, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                numeric = safe_float(cell.value)
                if numeric is None:
                    continue
                if -1.0 <= numeric <= 1.5:
                    cell.number_format = "0.00%"

        # 风险热力色阶
        for col_name in heatmap_cols_map.get(ws.title, []):
            col_idx = header_to_index.get(col_name)
            if not col_idx or max_row <= 1:
                continue
            col_letter = get_column_letter(col_idx)
            data_range = f"{col_letter}2:{col_letter}{max_row}"
            try:
                ws.conditional_formatting.add(
                    data_range,
                    ColorScaleRule(
                        start_type="min", start_color="FDE2E2",
                        mid_type="percentile", mid_value=50, mid_color="FFF4CC",
                        end_type="max", end_color="C6EFCE"
                    )
                )
            except Exception:
                pass

        # 列宽自适应
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, max_row + 1):
                text = clean_text(ws.cell(row=row_idx, column=col_idx).value)
                max_len = max(max_len, len(text))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 55)

    try:
        wb.save(file_path)
    except Exception as exc:
        logging.warning("Excel美化保存失败：%s | %s", file_path, exc)


def save_friendly_excel(file_path, sheet_frames, percent_cols_map=None, heatmap_cols_map=None):
    """将多张DataFrame写入一个Excel，并统一做美化。"""
    try:
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            for sheet_name, df_sheet in sheet_frames.items():
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
        style_excel_workbook(file_path, percent_cols_map=percent_cols_map, heatmap_cols_map=heatmap_cols_map)
    except Exception as exc:
        logging.warning("友好Excel导出失败，退回基础导出：%s | %s", file_path, exc)
        first_sheet_name = next(iter(sheet_frames.keys()))
        first_df = sheet_frames[first_sheet_name]
        first_df.to_excel(file_path, index=False)


# -----------------------
# 基础工具函数
# -----------------------
def safe_read_csv(path):
    """安全读取员工数据（兼容CSV/Excel）"""
    logging.info("读取员工数据: %s", path)
    if not os.path.exists(path):
        logging.error("员工数据文件不存在：%s", path)
        sys.exit(1)
    file_ext = os.path.splitext(path)[1].lower()
    if file_ext in {".xlsx", ".xls"}:
        return pd.read_excel(path)
    return pd.read_csv(path)


def safe_read_excel(path):
    """安全读取Excel文件"""
    logging.info("读取政策数据: %s", path)
    if not os.path.exists(path):
        logging.error("Excel文件不存在：%s", path)
        sys.exit(1)
    return pd.read_excel(path)


def onehot_encoder_compat(handle_unknown="ignore"):
    """兼容不同sklearn版本的OneHotEncoder"""
    try:
        return OneHotEncoder(handle_unknown=handle_unknown, sparse_output=False)
    except TypeError:
        return OneHotEncoder(handle_unknown=handle_unknown, sparse=False)


def ensure_lightgbm():
    """确保lightgbm已安装"""
    if lgb is None:
        logging.error("缺少 lightgbm，请先执行：pip install lightgbm")
        sys.exit(1)


def ensure_shap_warn():
    """检查shap，无则警告"""
    if shap is None:
        logging.warning("未检测到 shap，Top-3 风险驱动和 SHAP 图将被跳过。安装：pip install shap")
        return False
    return True


# -----------------------
# 政策识别与语义编码工具
# -----------------------
POLICY_COLUMN_CANDIDATES = {
    "title": ["文章标题", "标题", "title", "政策标题", "文件标题", "名称"],
    "content": ["正文内容", "内容", "body", "正文", "政策内容", "主要内容", "摘要"],
    "time": ["发布时间", "发布日期", "time", "date", "publish_date", "发文时间", "发布时间间"],
    "role": ["适用岗位", "岗位", "job_role", "岗位名称", "岗位类别", "职位", "适用对象", "适用群体"],
    "department": ["适用部门", "部门", "department", "所属部门", "适用业务条线"],
    "source": ["来源", "政策来源", "发布机构", "发布单位", "发文机关"]
}

JOB_ROLE_ALIAS_GROUPS = {
    "Sales Executive": ["sales executive", "salesexecutive", "销售主管", "销售经理", "销售专员", "销售执行", "销售顾问"],
    "Research Scientist": ["research scientist", "researchscientist", "研究科学家", "科研人员", "研发人员", "研究员"],
    "Laboratory Technician": ["laboratory technician", "laboratorytechnician", "实验室技术员", "检验技术员", "技术员", "实验员"],
    "Manufacturing Director": ["manufacturing director", "manufacturingdirector", "制造总监", "生产总监", "制造负责人", "生产负责人"],
    "Healthcare Representative": ["healthcare representative", "healthcarerepresentative", "医疗代表", "医药代表", "健康顾问"],
    "Manager": ["manager", "管理者", "经理", "主管"],
    "Sales Representative": ["sales representative", "salesrepresentative", "销售代表", "业务代表"],
    "Research Director": ["research director", "researchdirector", "研发总监", "研究总监", "科研总监"],
    "Human Resources": ["human resources", "humanresources", "hr", "人力资源", "人事"]
}

DEPARTMENT_ALIAS_GROUPS = {
    "Sales": ["sales", "销售", "市场销售"],
    "Research & Development": ["researchdevelopment", "research&development", "r&d", "研发", "研究开发", "技术研发", "科研"],
    "Human Resources": ["human resources", "humanresources", "hr", "人力资源", "人事"]
}

POLICY_TOPIC_RULES = {
    "compensation": {"label": "薪酬激励", "keywords": ["补贴", "津贴", "薪酬", "工资", "奖金", "绩效", "福利", "社保", "公积金", "激励"]},
    "development": {"label": "培训发展", "keywords": ["培训", "培养", "学习", "技能", "课程", "研修", "能力提升", "导师", "继续教育"]},
    "promotion": {"label": "晋升成长", "keywords": ["晋升", "职级", "成长", "发展通道", "人才梯队", "职业发展", "晋级", "骨干"]},
    "worklife": {"label": "工作生活", "keywords": ["弹性", "休假", "加班", "工时", "差旅", "平衡", "双休", "远程", "调休"]},
    "environment": {"label": "环境氛围", "keywords": ["环境", "文化", "氛围", "团队", "办公", "协同", "体验", "满意度"]},
    "recognition": {"label": "认可表彰", "keywords": ["表彰", "荣誉", "奖励", "评优", "认可", "先进", "嘉奖"]},
    "housing": {"label": "住房保障", "keywords": ["住房", "租房", "通勤", "落户", "安家", "交通", "宿舍", "购房"]},
    "care": {"label": "关怀保障", "keywords": ["健康", "医疗", "子女", "家庭", "托育", "保险", "心理", "关怀", "帮扶", "慰问"]}
}
POLICY_TOPIC_KEYS = list(POLICY_TOPIC_RULES.keys())
POSITIVE_POLICY_WORDS = ["支持", "补贴", "奖励", "扶持", "优惠", "资助", "鼓励", "提升", "保障", "优化"]
NEGATIVE_POLICY_WORDS = ["取消", "减少", "撤销", "处罚", "限制", "收紧", "压减", "叫停", "约束"]


def clean_text(value):
    """清洗文本，兼容空值和异常字符串"""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null", "nat"}:
        return ""
    return re.sub(r"\s+", " ", text)


def normalize_identifier(value):
    """统一文本标识，便于列名/岗位名匹配"""
    text = clean_text(value)
    if not text:
        return ""
    return re.sub(r"[^\u4e00-\u9fa5a-zA-Z0-9]+", "", text).lower()


ROLE_ALIAS_TO_CANONICAL = {
    normalize_identifier(alias): canonical
    for canonical, aliases in JOB_ROLE_ALIAS_GROUPS.items()
    for alias in ([canonical] + aliases)
}

DEPARTMENT_ALIAS_TO_CANONICAL = {
    normalize_identifier(alias): canonical
    for canonical, aliases in DEPARTMENT_ALIAS_GROUPS.items()
    for alias in ([canonical] + aliases)
}


def load_bert_model():
    """加载预训练BERT模型（中文适配）"""
    try:
        tokenizer = BertTokenizer.from_pretrained('bert-base-chinese')
        model = BertModel.from_pretrained('bert-base-chinese')
        model.eval()
        logging.info("✅ BERT模型加载成功")
        return tokenizer, model
    except Exception as e:
        logging.error("BERT模型加载失败：%s，使用TF-IDF替代", e)
        return None, None


def ensure_list(value):
    """将标量/列表统一转为干净的字符串列表"""
    if isinstance(value, list):
        return [clean_text(v) for v in value if clean_text(v)]
    if isinstance(value, tuple):
        return [clean_text(v) for v in value if clean_text(v)]
    text = clean_text(value)
    return [text] if text else []


def dedupe_keep_order(values):
    """去重并保留原有顺序"""
    seen = set()
    results = []
    for value in values:
        text = clean_text(value)
        if text and text not in seen:
            seen.add(text)
            results.append(text)
    return results


def get_text_series(df, col):
    """安全获取文本列，没有时返回空序列"""
    if col and col in df.columns:
        return df[col].apply(clean_text)
    return pd.Series([""] * len(df), index=df.index, dtype="object")


def find_first_existing_column(df, candidates):
    """兼容列名大小写/符号差异，自动找到第一个可用列"""
    normalized_map = {normalize_identifier(col): col for col in df.columns}
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
        normalized = normalize_identifier(candidate)
        if normalized in normalized_map:
            return normalized_map[normalized]
    return None


def split_multi_value_text(text):
    """切分多值字段，兼容中英文分隔符"""
    value = clean_text(text)
    if not value:
        return []
    parts = re.split(r"[，,、/|；;\n]+", value)
    parts = [clean_text(part) for part in parts if clean_text(part)]
    return parts if parts else [value]


def canonicalize_job_role(value):
    """将岗位名映射到统一标准岗位名"""
    text = clean_text(value)
    if not text:
        return ""
    return ROLE_ALIAS_TO_CANONICAL.get(normalize_identifier(text), text)


def job_role_to_key(value):
    """岗位标准键，供DataFrame merge使用"""
    return normalize_identifier(canonicalize_job_role(value))


def canonicalize_department(value):
    """将部门名映射到统一标准部门名"""
    text = clean_text(value)
    if not text:
        return ""
    return DEPARTMENT_ALIAS_TO_CANONICAL.get(normalize_identifier(text), text)


def department_to_key(value):
    """部门标准键，供规则匹配使用"""
    return normalize_identifier(canonicalize_department(value))


def match_aliases_from_text(text, alias_map):
    """从自由文本中回捞岗位/部门别名"""
    normalized_text = normalize_identifier(text)
    matches = []
    for alias_key, canonical in alias_map.items():
        if not alias_key:
            continue
        if len(alias_key) <= 2:
            if normalized_text == alias_key:
                matches.append(canonical)
        elif alias_key in normalized_text:
            matches.append(canonical)
    return dedupe_keep_order(matches)


def extract_targets_from_text(explicit_value, fallback_text, canonicalize_fn, key_fn, alias_map):
    """优先用结构化字段抽取目标对象，缺失时再从全文回捞"""
    labels = []
    for part in split_multi_value_text(explicit_value):
        canonical = canonicalize_fn(part)
        if canonical:
            labels.append(canonical)
    if not labels:
        labels = match_aliases_from_text(fallback_text, alias_map)
    labels = dedupe_keep_order(labels)
    keys = dedupe_keep_order([key_fn(label) for label in labels if key_fn(label)])
    return labels, keys


def extract_policy_targets(role_value="", department_value="", full_text=""):
    """抽取政策适用岗位和部门"""
    role_labels, role_keys = extract_targets_from_text(
        role_value, full_text, canonicalize_job_role, job_role_to_key, ROLE_ALIAS_TO_CANONICAL
    )
    department_labels, department_keys = extract_targets_from_text(
        department_value, full_text, canonicalize_department, department_to_key, DEPARTMENT_ALIAS_TO_CANONICAL
    )
    return pd.Series({
        "target_role_labels": role_labels,
        "target_role_keys": role_keys,
        "target_department_labels": department_labels,
        "target_department_keys": department_keys
    })


def detect_policy_topics(text):
    """基于关键词识别政策主题"""
    clean = clean_text(text)
    scores = {}
    for topic_key, config in POLICY_TOPIC_RULES.items():
        scores[topic_key] = float(sum(clean.count(keyword) for keyword in config["keywords"]))
    return scores


def topic_labels_from_scores(score_dict):
    """将主题得分转换为可读标签"""
    labels = [POLICY_TOPIC_RULES[key]["label"] for key in POLICY_TOPIC_KEYS if score_dict.get(key, 0.0) > 0]
    return labels if labels else ["综合支持"]


def topic_vector_from_scores(score_dict):
    """将主题得分转换为归一化向量"""
    vector = np.array([float(score_dict.get(key, 0.0)) for key in POLICY_TOPIC_KEYS], dtype=float)
    total = vector.sum()
    if total > 0:
        vector = vector / total
    return vector


def compute_policy_sentiment(text):
    """识别政策语气：支持型为正、约束型为负"""
    clean = clean_text(text)
    pos = sum(clean.count(word) for word in POSITIVE_POLICY_WORDS)
    neg = sum(clean.count(word) for word in NEGATIVE_POLICY_WORDS)
    if pos == 0 and neg == 0:
        return 0.0
    return float(np.clip((pos - neg) / (pos + neg + 1e-8), -1.0, 1.0))


def compute_time_weight(pub_time, newest, half_life_days):
    """按发布时间计算时间衰减权重"""
    try:
        if pd.isna(pub_time):
            return 0.5
        delta_days = max(0, (newest - pd.to_datetime(pub_time)).days)
        return float(0.5 ** (delta_days / max(half_life_days, 1)))
    except Exception:
        return 0.5


def build_text_embeddings(texts, tokenizer=None, model=None, batch_size=16, max_length=256):
    """统一文本向量化，BERT优先，失败回退到字符级TF-IDF"""
    cleaned_texts = [clean_text(text) or "空文本" for text in texts]
    if not cleaned_texts:
        return np.zeros((0, 1)), "empty"

    if tokenizer is not None and model is not None:
        try:
            device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
            model = model.to(device)
            model.eval()
            batches = []
            iterator = range(0, len(cleaned_texts), batch_size)
            if len(cleaned_texts) > batch_size:
                iterator = tqdm(iterator, desc="🔄 BERT编码", leave=False)
            for start in iterator:
                batch_texts = cleaned_texts[start:start + batch_size]
                inputs = tokenizer(
                    batch_texts,
                    return_tensors="pt",
                    truncation=True,
                    padding=True,
                    max_length=max_length
                )
                inputs = {k: v.to(device) for k, v in inputs.items()}
                with torch.no_grad():
                    outputs = model(**inputs)
                batches.append(outputs.last_hidden_state.mean(dim=1).cpu().numpy())
            return np.vstack(batches), "bert"
        except Exception as e:
            logging.warning("BERT批量编码失败：%s，回退TF-IDF", e)

    max_features = min(1024, max(128, len(cleaned_texts) * 8))
    vectorizer = TfidfVectorizer(max_features=max_features, analyzer="char", ngram_range=(2, 4))
    matrix = vectorizer.fit_transform(cleaned_texts).toarray()
    return matrix, "tfidf"


def text_to_embedding(text, tokenizer=None, model=None):
    """文本转嵌入向量（单条兼容包装）"""
    embeddings, _ = build_text_embeddings([text], tokenizer, model)
    return embeddings[0]


def is_prepared_policy_df(df):
    """判断政策DataFrame是否已经完成标准化处理"""
    required_cols = {
        "full_text", "semantic_text", "topic_vector", "time_weight",
        "target_role_keys", "target_department_keys", "policy_score", "embedding"
    }
    return isinstance(df, pd.DataFrame) and required_cols.issubset(df.columns)


def prepare_policy_dataframe(policy_source, tokenizer=None, model=None, half_life_days=TIME_DECAY_HALF_LIFE_DAYS):
    """统一的政策识别模块：列识别、岗位识别、主题识别、语义编码"""
    if isinstance(policy_source, pd.DataFrame):
        df = policy_source.copy()
        source_path = None
    else:
        source_path = policy_source
        if not os.path.exists(source_path):
            logging.warning("❌ 政策文件未找到：%s", source_path)
            return pd.DataFrame()
        df = safe_read_excel(source_path)

    if df.empty:
        logging.warning("❌ 政策数据为空，跳过政策识别")
        return pd.DataFrame()

    if is_prepared_policy_df(df):
        return df.copy()

    title_col = find_first_existing_column(df, POLICY_COLUMN_CANDIDATES["title"])
    content_col = find_first_existing_column(df, POLICY_COLUMN_CANDIDATES["content"])
    time_col = find_first_existing_column(df, POLICY_COLUMN_CANDIDATES["time"])
    role_col = find_first_existing_column(df, POLICY_COLUMN_CANDIDATES["role"])
    department_col = find_first_existing_column(df, POLICY_COLUMN_CANDIDATES["department"])
    source_col = find_first_existing_column(df, POLICY_COLUMN_CANDIDATES["source"])

    has_full_text = "full_text" in df.columns and df["full_text"].apply(clean_text).ne("").any()
    if not has_full_text and title_col is None and content_col is None:
        logging.warning("❌ 政策表中未找到可用文本列（标题/正文/full_text）")
        return pd.DataFrame()

    title_series = get_text_series(df, title_col)
    content_series = get_text_series(df, content_col)
    source_series = get_text_series(df, source_col)
    if has_full_text:
        full_text = df["full_text"].apply(clean_text)
    else:
        full_text = (title_series + " " + content_series).str.strip().apply(clean_text)

    df["policy_title"] = title_series
    df["policy_body"] = content_series
    df["policy_source"] = source_series
    df["full_text"] = full_text

    if time_col:
        df["publish_time"] = pd.to_datetime(df[time_col], errors="coerce")
    else:
        fallback_time = pd.Timestamp.fromtimestamp(os.path.getmtime(source_path)) if source_path and os.path.exists(source_path) else pd.Timestamp.now()
        df["publish_time"] = fallback_time

    newest = df["publish_time"].dropna().max()
    if pd.isna(newest):
        newest = pd.Timestamp.now()
    df["time_weight"] = df["publish_time"].apply(lambda x: compute_time_weight(x, newest, half_life_days))

    if {"target_role_labels", "target_role_keys", "target_department_labels", "target_department_keys"}.issubset(df.columns):
        for column in ["target_role_labels", "target_role_keys", "target_department_labels", "target_department_keys"]:
            df[column] = df[column].apply(ensure_list)
    else:
        target_info = df.apply(
            lambda row: extract_policy_targets(
                clean_text(row[role_col]) if role_col else "",
                clean_text(row[department_col]) if department_col else "",
                row["full_text"]
            ),
            axis=1
        )
        df = pd.concat([df, target_info], axis=1)

    df["topic_scores"] = df["full_text"].apply(detect_policy_topics)
    df["topic_vector"] = df["topic_scores"].apply(topic_vector_from_scores)
    df["policy_tags"] = df["topic_scores"].apply(lambda x: "、".join(topic_labels_from_scores(x)))
    df["matched_topic_count"] = df["topic_scores"].apply(lambda x: sum(1 for value in x.values() if value > 0))
    df["topic_hit_total"] = df["topic_scores"].apply(lambda x: float(sum(x.values())))
    df["topic_strength"] = np.clip((df["matched_topic_count"] * 0.6 + df["topic_hit_total"] * 0.4) / 4.0, 0, 1)
    df["policy_sentiment"] = df["full_text"].apply(compute_policy_sentiment)

    def build_semantic_text(row):
        pieces = [
            row["policy_title"],
            row["full_text"],
            row["policy_tags"],
            " ".join(ensure_list(row["target_role_labels"])),
            " ".join(ensure_list(row["target_department_labels"])),
            row["policy_source"]
        ]
        return clean_text(" ".join([piece for piece in pieces if clean_text(piece)]))

    df["semantic_text"] = df.apply(build_semantic_text, axis=1)

    try:
        tfidf = TfidfVectorizer(max_features=2000, analyzer="char", ngram_range=(2, 4))
        tfidf_matrix = tfidf.fit_transform(df["semantic_text"].replace("", "空文本"))
        doc_sum = tfidf_matrix.sum(axis=1).A1
        if np.isclose(doc_sum.max(), doc_sum.min()):
            policy_hotness = np.full(len(df), 0.5)
        else:
            policy_hotness = (doc_sum - doc_sum.min()) / (doc_sum.max() - doc_sum.min() + 1e-8)
    except Exception:
        policy_hotness = np.full(len(df), 0.5)
    df["policy_hotness"] = policy_hotness

    embeddings, embedding_backend = build_text_embeddings(df["semantic_text"].tolist(), tokenizer, model)
    df["embedding"] = [embeddings[i] for i in range(len(df))]
    df["embedding_backend"] = embedding_backend

    if len(df) == 1:
        semantic_score = np.ones(1)
    else:
        semantic_center = embeddings.mean(axis=0, keepdims=True)
        semantic_score = np.asarray(calculate_similarity(embeddings, semantic_center)).reshape(-1)
        semantic_score = np.clip((semantic_score + 1.0) / 2.0, 0, 1)
    df["semantic_score"] = semantic_score

    support_bias = np.clip((df["policy_sentiment"] + 1.0) / 2.0, 0, 1)
    df["policy_score"] = (
        100 * (
            0.28 * df["policy_hotness"]
            + 0.24 * df["semantic_score"]
            + 0.22 * df["time_weight"]
            + 0.16 * df["topic_strength"]
            + 0.10 * support_bias
        )
    ).clip(0, 100)

    logging.info("✅ 政策识别模块完成：识别到 %s 条政策，语义编码模式=%s", len(df), embedding_backend)
    return df


def calculate_similarity(vec1, vec2):
    """计算向量余弦相似度，兼容单向量和矩阵"""
    arr1 = np.atleast_2d(np.asarray(vec1, dtype=float))
    arr2 = np.atleast_2d(np.asarray(vec2, dtype=float))
    arr1 = arr1 / (np.linalg.norm(arr1, axis=1, keepdims=True) + 1e-8)
    arr2 = arr2 / (np.linalg.norm(arr2, axis=1, keepdims=True) + 1e-8)
    sim = arr1 @ arr2.T
    if np.asarray(vec1).ndim == 1 and np.asarray(vec2).ndim == 1:
        return float(sim[0, 0])
    return sim


# -----------------------
# 新增：可视化工具函数（4类核心图表）
# -----------------------
def plot_model_metrics(metrics, save_name="model_metrics.png", alias_names=None):
    """可视化模型性能（训练/验证/测试集AUC/Accuracy/F1对比）"""
    save_path = os.path.join(CURRENT_DIR, save_name)
    alias_names = alias_names or ["model-metric.png"]
    try:
        metric_plan = [
            ("auc", "AUC"),
            ("acc", "Accuracy"),
            ("f1", "F1"),
        ]
        split_plan = [
            ("train", "训练集"),
            ("valid", "验证集"),
            ("test", "测试集"),
        ]

        rows = []
        for split_key, split_label in split_plan:
            for metric_key, metric_label in metric_plan:
                full_key = f"{split_key}_{metric_key}"
                numeric_value = safe_float(metrics.get(full_key))
                if numeric_value is None:
                    continue
                rows.append({
                    "数据集": split_label,
                    "评估指标": metric_label,
                    "Score": numeric_value
                })

        if not rows:
            logging.warning("❌ model metrics图未生成：未找到 train/valid/test + auc/acc/f1 指标")
            return

        metrics_df = pd.DataFrame(rows)

        plt.figure(figsize=(10, 6))
        ax = sns.barplot(
            x="评估指标",
            y="Score",
            hue="数据集",
            data=metrics_df,
            order=[item[1] for item in metric_plan],
            hue_order=[item[1] for item in split_plan],
            palette="Set2",
            edgecolor="black"
        )
        ax.set_title("Model Metrics (Train / Valid / Test)", fontsize=14, fontweight="bold", pad=18)
        ax.set_xlabel("Metric", fontsize=12)
        ax.set_ylabel("Score", fontsize=12)
        ax.set_ylim(0, 1.08)

        # 添加柱子上的数值标签
        for patch in ax.patches:
            height = patch.get_height()
            if pd.isna(height):
                continue
            ax.annotate(
                f"{height:.3f}",
                (patch.get_x() + patch.get_width() / 2, height + 0.01),
                ha="center", va="bottom", fontsize=9
            )

        legend = ax.legend(title="数据集", loc="upper right")
        if legend is None:
            logging.warning("⚠️ model metrics图例生成失败，但不影响图表输出")

        plt.tight_layout()
        plt.savefig(save_path, dpi=150, bbox_inches="tight")
        plt.close()
        logging.info("✅ 模型性能图已保存：%s", save_path)

        # 兼容额外文件名（例如 model-metric.png）
        for alias_name in alias_names:
            alias_path = os.path.join(CURRENT_DIR, alias_name)
            if os.path.abspath(alias_path) == os.path.abspath(save_path):
                continue
            try:
                shutil.copyfile(save_path, alias_path)
                logging.info("✅ 模型性能图别名已保存：%s", alias_path)
            except Exception as exc:
                logging.warning("⚠️ 模型性能图别名保存失败：%s | %s", alias_path, exc)
    except Exception as exc:
        plt.close("all")
        logging.warning("❌ 模型性能图生成失败：%s", exc)


def plot_feature_importance(model, preprocessor, top_n=15, save_name="feature_importance.png"):
    """可视化Top-N特征重要性（基于LightGBM）"""
    save_path = os.path.join(CURRENT_DIR, save_name)
    try:
        # 获取LightGBM子模型
        lgb_model = model.named_estimators_['lgb']

        # 提取特征名（处理数值+类别特征）
        feature_names = []
        for name, trans, cols in preprocessor.transformers_:
            if name == 'num':
                feature_names.extend(cols)
            elif name == 'cat':
                ohe = trans.named_steps['ohe']
                feature_names.extend(ohe.get_feature_names_out(cols))

        # 计算并排序特征重要性
        importances = lgb_model.feature_importances_
        feat_imp = pd.DataFrame({'特征名称': feature_names, '重要性得分': importances})
        feat_imp = feat_imp.sort_values('重要性得分', ascending=False).head(top_n)

        # 绘制水平柱状图（便于查看长特征名）
        plt.figure(figsize=(12, 8))
        sns.barplot(x='重要性得分', y='特征名称', data=feat_imp, palette='viridis', edgecolor='black')
        plt.title(f'Top-{top_n} Feature Importance', fontsize=14, fontweight='bold', pad=20)
        plt.xlabel('Importance Score', fontsize=12)
        plt.ylabel('Feature Name', fontsize=12)
        plt.tight_layout()
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        plt.close()
        logging.info(f"✅ 特征重要性图已保存：{save_path}")
    except Exception as e:
        logging.warning(f"❌ 特征重要性可视化失败：{e}")


def plot_attrition_risk_distribution(y_pred_prob, threshold=0.5, save_name="attrition_risk_distribution.png"):
    """可视化员工流失风险概率分布"""
    save_path = os.path.join(CURRENT_DIR, save_name)
    plt.figure(figsize=(10, 6))

    # 绘制直方图+核密度曲线
    sns.histplot(y_pred_prob, bins=30, kde=True, color='orange', alpha=0.7, edgecolor='black')
    # 添加风险阈值线
    if isinstance(threshold, dict) and threshold.get("type") == "segment":
        high_threshold = float(threshold.get("high_risk", threshold.get("base_threshold", 0.5)))
        standard_threshold = float(threshold.get("standard", threshold.get("base_threshold", 0.5)))
        plt.axvline(x=high_threshold, color='red', linestyle='--', linewidth=2, label=f'高风险组阈值（{high_threshold:.2f}）')
        plt.axvline(x=standard_threshold, color='blue', linestyle='--', linewidth=2, label=f'常规组阈值（{standard_threshold:.2f}）')
    else:
        threshold_value = float(threshold)
        plt.axvline(x=threshold_value, color='red', linestyle='--', linewidth=2, label=f'风险阈值（{threshold_value:.2f}）')

    plt.title('Employee Attrition Risk Probability Distribution', fontsize=14, fontweight='bold', pad=20)
    plt.xlabel('Attrition Probability', fontsize=12)
    plt.ylabel('Number of Employees', fontsize=12)
    plt.legend(fontsize=11)
    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches='tight')
    plt.close()
    logging.info(f"✅ 风险分布直方图已保存：{save_path}")


def plot_attrition_decision_view(df_pred, threshold=0.5, save_name="attrition_decision_view.png", alias_names=None):
    """可视化员工去留预测结果（排序风险曲线 + 去留占比）。"""
    alias_names = alias_names or []
    save_path = os.path.join(CURRENT_DIR, save_name)

    if df_pred is None or df_pred.empty or "流失概率" not in df_pred.columns or "预测流失标签" not in df_pred.columns:
        logging.warning("❌ 去留预测可视化失败：缺少必要字段（流失概率/预测流失标签）")
        return

    try:
        view_df = df_pred.copy()
        view_df["流失概率"] = pd.to_numeric(view_df["流失概率"], errors="coerce")
        view_df["预测流失标签"] = pd.to_numeric(view_df["预测流失标签"], errors="coerce").fillna(0).astype(int)
        view_df = view_df.dropna(subset=["流失概率"]).sort_values("流失概率", ascending=False).reset_index(drop=True)
        if view_df.empty:
            logging.warning("❌ 去留预测可视化失败：流失概率列全为空")
            return

        total = len(view_df)
        x = np.arange(1, total + 1)
        y = view_df["流失概率"].values
        colors = np.where(view_df["预测流失标签"].values == 1, "#d84b48", "#2f7ed8")

        plt.figure(figsize=(13, 6.8))
        gs = plt.GridSpec(1, 2, width_ratios=[2.4, 1.1])

        # 左图：风险排序曲线
        ax1 = plt.subplot(gs[0, 0])
        ax1.plot(x, y, color="#1f6feb", linewidth=1.1, alpha=0.65)
        ax1.scatter(x, y, c=colors, s=20, alpha=0.9, edgecolors="white", linewidth=0.25)

        if isinstance(threshold, dict):
            high_t = safe_float(threshold.get("high_risk", threshold.get("base_threshold", 0.5)))
            std_t = safe_float(threshold.get("standard", threshold.get("base_threshold", 0.5)))
            if high_t is not None:
                ax1.axhline(y=high_t, color="#d84b48", linestyle="--", linewidth=1.6, label=f"高风险阈值 {high_t:.2f}")
            if std_t is not None:
                ax1.axhline(y=std_t, color="#2962cc", linestyle="--", linewidth=1.4, label=f"常规阈值 {std_t:.2f}")
        else:
            t = safe_float(threshold)
            if t is not None:
                ax1.axhline(y=t, color="#d84b48", linestyle="--", linewidth=1.6, label=f"阈值 {t:.2f}")

        ax1.set_title("员工去留风险排序（按流失概率降序）", fontsize=13, fontweight="bold", pad=12)
        ax1.set_xlabel("员工排名（1=风险最高）", fontsize=11)
        ax1.set_ylabel("流失概率", fontsize=11)
        ax1.set_ylim(0, 1.05)
        ax1.grid(alpha=0.25)
        ax1.legend(loc="upper right", fontsize=9)

        # 右图：去留占比
        ax2 = plt.subplot(gs[0, 1])
        leave_count = int((view_df["预测流失标签"] == 1).sum())
        stay_count = int(total - leave_count)
        shares = [stay_count / total, leave_count / total]
        bars = ax2.bar(
            ["预测稳定", "预测流失"],
            [stay_count, leave_count],
            color=["#2f7ed8", "#d84b48"],
            edgecolor="black",
            linewidth=0.4
        )
        ax2.set_title("去留人数占比", fontsize=13, fontweight="bold", pad=12)
        ax2.set_ylabel("人数", fontsize=11)
        ax2.set_ylim(0, max(stay_count, leave_count, 1) * 1.25)
        for idx, bar in enumerate(bars):
            height = bar.get_height()
            ax2.text(
                bar.get_x() + bar.get_width() / 2,
                height + max(total * 0.008, 1),
                f"{int(height)}人\n{shares[idx]:.1%}",
                ha="center",
                va="bottom",
                fontsize=10
            )

        summary_text = (
            f"总人数: {total}\n"
            f"预测流失: {leave_count} ({leave_count / total:.1%})\n"
            f"预测稳定: {stay_count} ({stay_count / total:.1%})"
        )
        ax2.text(
            0.02,
            0.98,
            summary_text,
            transform=ax2.transAxes,
            fontsize=9,
            va="top",
            ha="left",
            bbox=dict(facecolor="#f5f9ff", edgecolor="#d6e5fb", boxstyle="round,pad=0.35")
        )

        plt.tight_layout()
        plt.savefig(save_path, dpi=150, bbox_inches="tight")
        plt.close()
        logging.info("✅ 员工去留预测可视化图已保存：%s", save_path)

        for alias_name in alias_names:
            alias_path = os.path.join(CURRENT_DIR, alias_name)
            if os.path.abspath(alias_path) == os.path.abspath(save_path):
                continue
            try:
                shutil.copyfile(save_path, alias_path)
                logging.info("✅ 员工去留预测可视化图别名已保存：%s", alias_path)
            except Exception as exc:
                logging.warning("⚠️ 去留预测可视化别名保存失败：%s | %s", alias_path, exc)
    except Exception as exc:
        plt.close("all")
        logging.warning("❌ 员工去留预测可视化失败：%s", exc)


def plot_policy_job_matching(policy_post_mapping, save_name="policy_job_matching.png"):
    """可视化政策-岗位匹配得分"""
    if not policy_post_mapping:
        logging.warning("❌ 无政策-岗位匹配数据，跳过该可视化")
        return

    save_path = os.path.join(CURRENT_DIR, save_name)
    # 整理匹配数据
    match_df = pd.DataFrame(list(policy_post_mapping.items()), columns=['岗位名称', '政策匹配得分'])
    match_df = match_df.sort_values('政策匹配得分', ascending=False)

    # 绘制水平柱状图
    plt.figure(figsize=(12, 6))
    sns.barplot(x='政策匹配得分', y='岗位名称', data=match_df, palette='coolwarm', edgecolor='black')
    plt.title('Policy-Job Matching Score by Role', fontsize=14, fontweight='bold', pad=20)
    plt.xlabel('Policy Matching Score', fontsize=12)
    plt.ylabel('Job Role', fontsize=12)
    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches='tight')
    plt.close()
    logging.info(f"✅ 政策-岗位匹配图已保存：{save_path}")


# -----------------------
# 员工数据读取与预处理
# -----------------------
def load_and_preprocess_employee(path):
    """加载员工数据并预处理（缺失值填充、特征衍生）"""
    df = safe_read_csv(path)

    # 删除无用列
    for c in DROP_COLS:
        if c in df.columns:
            df.drop(columns=c, inplace=True)

    # 标签编码（Attrition→0/1）
    if "Attrition" in df.columns:
        df["AttritionFlag"] = df["Attrition"].map({"Yes": 1, "No": 0})

    # 缺失值填充（数值型→中位数，类别型→Missing）
    num_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    cat_cols = df.select_dtypes(include=["object"]).columns.tolist()
    df[num_cols] = df[num_cols].fillna(df[num_cols].median())
    df[cat_cols] = df[cat_cols].fillna("Missing")

    # 衍生特征：收入对数
    if "MonthlyIncome" in df.columns:
        df["MonthlyIncome_log"] = np.log1p(df["MonthlyIncome"])

    logging.info(f"✅ 员工数据预处理完成，数据形状：{df.shape}")
    return df


# -----------------------
# 交互特征生成（增强模型预测能力）
# -----------------------
def add_interaction_features(df):
    """生成更贴合离职场景的员工画像交互特征"""
    if 'OverTime' in df.columns:
        df['OverTimeFlag'] = df['OverTime'].map({'No': 0, 'Yes': 1}).fillna(0).astype(float)

    if 'MonthlyIncome' in df.columns and 'YearsAtCompany' in df.columns:
        df['Income_per_YearAtCompany'] = df['MonthlyIncome'] / (df['YearsAtCompany'] + 1)  # 避免除0
    if 'MonthlyIncome' in df.columns and 'TotalWorkingYears' in df.columns:
        df['Income_per_WorkYear'] = df['MonthlyIncome'] / (df['TotalWorkingYears'] + 1)
    if 'MonthlyIncome' in df.columns and 'JobLevel' in df.columns:
        df['Income_per_JobLevel'] = df['MonthlyIncome'] / (df['JobLevel'] + 1)
    if 'Age' in df.columns and 'WorkLifeBalance' in df.columns:
        df['Age_WorkBalance'] = df['Age'] * df['WorkLifeBalance']  # 年龄×工作生活平衡
    if 'BusinessTravel' in df.columns:
        df['TravelRisk'] = df['BusinessTravel'].map(
            {'Non-Travel': 0, 'Travel_Rarely': 1, 'Travel_Frequently': 2}).fillna(0).astype(float)  # 出差风险

    satisfaction_cols = [
        col for col in [
            'JobSatisfaction', 'EnvironmentSatisfaction',
            'RelationshipSatisfaction', 'WorkLifeBalance'
        ] if col in df.columns
    ]
    if satisfaction_cols:
        df['SatisfactionIndex'] = df[satisfaction_cols].mean(axis=1)
        if len(satisfaction_cols) > 1:
            df['SatisfactionGap'] = df[satisfaction_cols].max(axis=1) - df[satisfaction_cols].min(axis=1)

    if 'JobInvolvement' in df.columns and 'WorkLifeBalance' in df.columns:
        df['EngagementBalanceScore'] = df['JobInvolvement'] * df['WorkLifeBalance']

    if 'YearsSinceLastPromotion' in df.columns and 'YearsAtCompany' in df.columns:
        df['PromotionWaitRatio'] = df['YearsSinceLastPromotion'] / (df['YearsAtCompany'] + 1)
    if 'YearsInCurrentRole' in df.columns and 'YearsAtCompany' in df.columns:
        df['RoleStagnationRatio'] = df['YearsInCurrentRole'] / (df['YearsAtCompany'] + 1)
    if 'YearsWithCurrManager' in df.columns and 'YearsAtCompany' in df.columns:
        df['ManagerTenureRatio'] = df['YearsWithCurrManager'] / (df['YearsAtCompany'] + 1)
    if 'YearsInCurrentRole' in df.columns and 'YearsSinceLastPromotion' in df.columns:
        df['PromotionRoleGap'] = np.maximum(df['YearsInCurrentRole'] - df['YearsSinceLastPromotion'], 0)

    if 'NumCompaniesWorked' in df.columns and 'TotalWorkingYears' in df.columns:
        df['ExternalMobilityRatio'] = df['NumCompaniesWorked'] / (df['TotalWorkingYears'] + 1)

    if 'DistanceFromHome' in df.columns:
        if 'OverTimeFlag' in df.columns:
            df['DistanceOverTimePressure'] = df['DistanceFromHome'] * (1 + df['OverTimeFlag'])
        else:
            df['DistanceOverTimePressure'] = df['DistanceFromHome']

    if 'TravelRisk' in df.columns and 'OverTimeFlag' in df.columns:
        df['TravelOverTimeRisk'] = df['TravelRisk'] * (1 + df['OverTimeFlag'])

    if 'StockOptionLevel' in df.columns:
        df['LowStockOptionFlag'] = (df['StockOptionLevel'] <= 0).astype(float)
        if 'OverTimeFlag' in df.columns:
            df['LowStockOverTimeRisk'] = df['LowStockOptionFlag'] * df['OverTimeFlag']

    stress_flag_cols = []
    if 'WorkLifeBalance' in df.columns:
        df['LowWorkLifeFlag'] = (df['WorkLifeBalance'] <= 2).astype(float)
        stress_flag_cols.append('LowWorkLifeFlag')
    if 'JobSatisfaction' in df.columns:
        df['LowJobSatisfactionFlag'] = (df['JobSatisfaction'] <= 2).astype(float)
        stress_flag_cols.append('LowJobSatisfactionFlag')
    if 'EnvironmentSatisfaction' in df.columns:
        df['LowEnvironmentSatisfactionFlag'] = (df['EnvironmentSatisfaction'] <= 2).astype(float)
        stress_flag_cols.append('LowEnvironmentSatisfactionFlag')
    if 'OverTimeFlag' in df.columns:
        stress_flag_cols.append('OverTimeFlag')
    if stress_flag_cols:
        df['StressLoadScore'] = df[stress_flag_cols].sum(axis=1)

    if 'YearsAtCompany' in df.columns:
        df['TenureBand'] = pd.cut(
            df['YearsAtCompany'],
            bins=[-np.inf, 2, 5, 10, np.inf],
            labels=['0-2年', '3-5年', '6-10年', '10年以上']
        ).astype(str).replace('nan', 'Missing')
    if 'DistanceFromHome' in df.columns:
        df['CommuteBand'] = pd.cut(
            df['DistanceFromHome'],
            bins=[-np.inf, 5, 15, np.inf],
            labels=['近距离', '中距离', '远距离']
        ).astype(str).replace('nan', 'Missing')
    if 'YearsSinceLastPromotion' in df.columns:
        df['PromotionWaitBand'] = pd.cut(
            df['YearsSinceLastPromotion'],
            bins=[-np.inf, 1, 3, np.inf],
            labels=['近期晋升', '观察期', '长期未晋升']
        ).astype(str).replace('nan', 'Missing')

    logging.info("✅ 交互特征生成完成")
    return df


# -----------------------
# 构建增强宏观政策指数
# -----------------------
def build_policy_macro_index_enhanced(path, tokenizer=None, model=None, half_life_days=180):
    """构建岗位级/全局宏观政策指数（复用统一政策识别模块）"""
    policy_df = path.copy() if is_prepared_policy_df(path) else prepare_policy_dataframe(path, tokenizer, model, half_life_days)
    if policy_df.empty:
        return pd.DataFrame()

    expanded_rows = []
    for _, row in policy_df.iterrows():
        role_labels = ensure_list(row.get("target_role_labels", []))
        role_keys = ensure_list(row.get("target_role_keys", []))
        pair_count = min(len(role_labels), len(role_keys))
        for idx in range(pair_count):
            if role_keys[idx]:
                expanded_rows.append({
                    "JobRole": role_labels[idx],
                    "JobRoleKey": role_keys[idx],
                    "doc_policy_score": float(row["policy_score"])
                })

    if expanded_rows:
        expanded_df = pd.DataFrame(expanded_rows)
        grouped = expanded_df.groupby(["JobRoleKey", "JobRole"])["doc_policy_score"].agg(["mean", "count"]).reset_index()
        max_count = max(int(grouped["count"].max()), 1)
        grouped["macro_index"] = (grouped["mean"] * 0.85 + grouped["count"] / max_count * 15).clip(0, 100)
        logging.info("✅ 构建岗位级宏观政策指数完成，覆盖岗位数：%s", len(grouped))
        return grouped[["JobRoleKey", "JobRole", "macro_index"]]

    macro_index_value = float(policy_df["policy_score"].mean())
    logging.info("✅ 构建全局宏观政策指数完成，指数值=%.2f", macro_index_value)
    return pd.DataFrame({"macro_index": [macro_index_value]})


# -----------------------
# 政策-员工语义匹配（增强特征）
# -----------------------
def compute_policy_impact(policy_path, tokenizer, model):
    """计算政策影响得分与岗位匹配字典"""
    df_policy = policy_path.copy() if is_prepared_policy_df(policy_path) else prepare_policy_dataframe(policy_path, tokenizer, model)
    if df_policy.empty:
        return 0.0, {}, []

    policy_post_mapping = {}
    for _, row in df_policy.iterrows():
        role_labels = ensure_list(row.get("target_role_labels", []))
        if not role_labels:
            continue
        for role_label in role_labels:
            policy_post_mapping.setdefault(role_label, []).append(float(row["policy_score"]))

    policy_post_mapping = {
        role: float(np.mean(scores))
        for role, scores in policy_post_mapping.items() if scores
    }
    total_policy_score = float(df_policy["policy_score"].mean())
    logging.info("📊 综合政策影响总分 = %.3f", total_policy_score)
    return total_policy_score, policy_post_mapping, df_policy["embedding"].tolist()


def add_policy_effect(df_emp, policy_df, tokenizer, model):
    """为员工添加政策识别和语义匹配特征"""
    df_emp = df_emp.copy()
    policy_df = policy_df.copy() if is_prepared_policy_df(policy_df) else prepare_policy_dataframe(policy_df, tokenizer, model)

    feature_cols = [
        "policy_match_mean", "policy_match_max", "policy_match_top3_mean", "policy_role_match_mean",
        "policy_support_score", "policy_constraint_score", "policy_net_support"
    ] + [f"policy_{topic_key}_exposure" for topic_key in POLICY_TOPIC_KEYS]

    if "JobRole" in df_emp.columns:
        df_emp["JobRoleKey"] = df_emp["JobRole"].apply(job_role_to_key)
    else:
        df_emp["JobRoleKey"] = ""
    if "Department" in df_emp.columns:
        df_emp["DepartmentKey"] = df_emp["Department"].apply(department_to_key)
    else:
        df_emp["DepartmentKey"] = ""

    if policy_df.empty:
        for col in feature_cols:
            df_emp[col] = 0.0
        logging.warning("❌ 政策数据为空，政策语义特征已全部置零")
        return df_emp

    stats = {
        "monthly_income_median": df_emp["MonthlyIncome"].median() if "MonthlyIncome" in df_emp.columns else np.nan,
        "distance_q75": df_emp["DistanceFromHome"].quantile(0.75) if "DistanceFromHome" in df_emp.columns else np.nan
    }

    def build_employee_policy_profile(row):
        scores = {key: 0.0 for key in POLICY_TOPIC_KEYS}
        fragments = []

        role = clean_text(row.get("JobRole", ""))
        department = clean_text(row.get("Department", ""))
        education = clean_text(row.get("EducationField", ""))

        if role:
            fragments.append(f"岗位 {role}")
        if department:
            fragments.append(f"部门 {department}")
        if education:
            fragments.append(f"专业 {education}")

        overtime = clean_text(row.get("OverTime", ""))
        if overtime == "Yes":
            scores["worklife"] += 1.4
            scores["care"] += 0.6
            fragments.append("关注加班治理 弹性休假 健康关怀")

        travel = clean_text(row.get("BusinessTravel", ""))
        if travel == "Travel_Frequently":
            scores["worklife"] += 1.0
            scores["compensation"] += 0.5
            fragments.append("关注差旅补贴 工作生活平衡")
        elif travel == "Travel_Rarely":
            scores["worklife"] += 0.4

        work_life_balance = row.get("WorkLifeBalance")
        if pd.notna(work_life_balance) and float(work_life_balance) <= 2:
            scores["worklife"] += 1.4
            scores["care"] += 0.4
            fragments.append("关注工作生活平衡 心理健康 员工关怀")

        monthly_income = row.get("MonthlyIncome")
        if pd.notna(monthly_income) and pd.notna(stats["monthly_income_median"]) and float(monthly_income) <= float(stats["monthly_income_median"]):
            scores["compensation"] += 1.3
            fragments.append("关注薪酬补贴 福利激励")

        years_since_last_promotion = row.get("YearsSinceLastPromotion")
        if pd.notna(years_since_last_promotion) and float(years_since_last_promotion) >= 3:
            scores["promotion"] += 1.3
            scores["development"] += 0.7
            fragments.append("关注晋升发展 职级成长")

        training_times = row.get("TrainingTimesLastYear")
        if pd.notna(training_times) and float(training_times) <= 1:
            scores["development"] += 1.1
            fragments.append("关注培训学习 技能提升")

        environment_satisfaction = row.get("EnvironmentSatisfaction")
        relationship_satisfaction = row.get("RelationshipSatisfaction")
        if (
            pd.notna(environment_satisfaction) and float(environment_satisfaction) <= 2
        ) or (
            pd.notna(relationship_satisfaction) and float(relationship_satisfaction) <= 2
        ):
            scores["environment"] += 1.2
            scores["care"] += 0.5
            fragments.append("关注工作环境 团队氛围 员工关怀")

        distance_from_home = row.get("DistanceFromHome")
        if pd.notna(distance_from_home) and pd.notna(stats["distance_q75"]) and float(distance_from_home) >= float(stats["distance_q75"]):
            scores["housing"] += 1.0
            fragments.append("关注住房交通 通勤补贴")

        job_level = row.get("JobLevel")
        total_working_years = row.get("TotalWorkingYears")
        if (
            pd.notna(job_level) and float(job_level) <= 2
            and pd.notna(total_working_years) and float(total_working_years) <= 5
        ):
            scores["development"] += 0.8
            scores["promotion"] += 0.4
            fragments.append("关注青年人才培养 职业发展")

        job_satisfaction = row.get("JobSatisfaction")
        if pd.notna(job_satisfaction) and float(job_satisfaction) <= 2:
            scores["recognition"] += 0.8
            scores["environment"] += 0.6
            fragments.append("关注认可激励 荣誉表彰")

        need_labels = topic_labels_from_scores(scores)
        fragments.append("重点需求 " + " ".join(need_labels))
        return pd.Series({
            "employee_policy_text": clean_text("；".join(dedupe_keep_order(fragments))),
            "employee_need_vector": topic_vector_from_scores(scores),
            "employee_need_tags": "、".join(need_labels)
        })

    employee_profile_df = df_emp.apply(build_employee_policy_profile, axis=1)
    df_emp = pd.concat([df_emp, employee_profile_df], axis=1)

    all_texts = policy_df["semantic_text"].tolist() + df_emp["employee_policy_text"].tolist()
    all_embeddings, embedding_backend = build_text_embeddings(all_texts, tokenizer, model)
    policy_embeddings = all_embeddings[:len(policy_df)]
    employee_embeddings = all_embeddings[len(policy_df):]

    semantic_similarity = calculate_similarity(employee_embeddings, policy_embeddings)
    semantic_similarity = np.clip((semantic_similarity + 1.0) / 2.0, 0, 1)

    policy_topic_matrix = np.vstack([np.asarray(vec, dtype=float) for vec in policy_df["topic_vector"]])
    employee_need_matrix = np.vstack([np.asarray(vec, dtype=float) for vec in df_emp["employee_need_vector"]])
    topic_similarity = calculate_similarity(employee_need_matrix, policy_topic_matrix)
    topic_similarity = np.clip(topic_similarity, 0, 1)

    match_matrix = 0.7 * semantic_similarity + 0.3 * topic_similarity
    policy_strength = np.clip(policy_df["policy_score"].to_numpy(dtype=float) / 100.0, 0.05, 1.0)
    time_weight = np.clip(policy_df["time_weight"].to_numpy(dtype=float), 0.1, 1.0)
    match_matrix = match_matrix * policy_strength[np.newaxis, :] * time_weight[np.newaxis, :]

    emp_role_keys = df_emp["JobRoleKey"].fillna("").astype(str).tolist()
    emp_department_keys = df_emp["DepartmentKey"].fillna("").astype(str).tolist()
    policy_role_keys_list = [ensure_list(value) for value in policy_df["target_role_keys"]]
    policy_department_keys_list = [ensure_list(value) for value in policy_df["target_department_keys"]]

    role_bonus = np.ones_like(match_matrix)
    department_bonus = np.ones_like(match_matrix)
    for col_idx, role_keys in enumerate(policy_role_keys_list):
        if role_keys:
            role_bonus[:, col_idx] = np.array([1.20 if emp_key and emp_key in role_keys else 0.88 for emp_key in emp_role_keys], dtype=float)
    for col_idx, department_keys in enumerate(policy_department_keys_list):
        if department_keys:
            department_bonus[:, col_idx] = np.array(
                [1.10 if emp_key and emp_key in department_keys else 0.93 for emp_key in emp_department_keys],
                dtype=float
            )
    match_matrix = np.clip(match_matrix * role_bonus * department_bonus, 0, None)

    top_k = min(3, match_matrix.shape[1])
    sorted_scores = np.sort(match_matrix, axis=1)
    df_emp["policy_match_mean"] = match_matrix.mean(axis=1)
    df_emp["policy_match_max"] = match_matrix.max(axis=1)
    df_emp["policy_match_top3_mean"] = sorted_scores[:, -top_k:].mean(axis=1) if top_k else 0.0

    role_specific_scores = []
    for row_idx, emp_role_key in enumerate(emp_role_keys):
        role_mask = np.array([bool(role_keys) and emp_role_key in role_keys for role_keys in policy_role_keys_list], dtype=bool)
        role_specific_scores.append(match_matrix[row_idx, role_mask].mean() if role_mask.any() else df_emp["policy_match_mean"].iloc[row_idx])
    df_emp["policy_role_match_mean"] = role_specific_scores

    policy_sentiments = policy_df["policy_sentiment"].to_numpy(dtype=float)
    support_weights = np.clip(policy_sentiments, 0, None)
    constraint_weights = np.clip(-policy_sentiments, 0, None)
    if support_weights.sum() > 0:
        df_emp["policy_support_score"] = (match_matrix * support_weights[np.newaxis, :]).sum(axis=1) / (support_weights.sum() + 1e-8)
    else:
        df_emp["policy_support_score"] = 0.0
    if constraint_weights.sum() > 0:
        df_emp["policy_constraint_score"] = (match_matrix * constraint_weights[np.newaxis, :]).sum(axis=1) / (constraint_weights.sum() + 1e-8)
    else:
        df_emp["policy_constraint_score"] = 0.0
    df_emp["policy_net_support"] = df_emp["policy_support_score"] - df_emp["policy_constraint_score"]

    topic_exposure = match_matrix @ policy_topic_matrix
    topic_exposure = topic_exposure / (match_matrix.sum(axis=1, keepdims=True) + 1e-8)
    for topic_idx, topic_key in enumerate(POLICY_TOPIC_KEYS):
        df_emp[f"policy_{topic_key}_exposure"] = topic_exposure[:, topic_idx]

    df_emp.drop(columns=["employee_policy_text", "employee_need_vector", "employee_need_tags"], inplace=True, errors="ignore")
    logging.info("✅ 已为 %s 位员工添加政策识别/语义匹配特征，编码模式=%s", len(df_emp), embedding_backend)
    return df_emp


# -----------------------
# 构建数据预处理器（数值+类别特征）
# -----------------------
def build_preprocessor(df, numeric_override=None, categorical_override=None):
    """构建ColumnTransformer预处理器（数值特征标准化，类别特征One-Hot）"""
    # 自动识别数值/类别特征（或使用自定义列表）
    if numeric_override is None:
        num_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        num_cols = [c for c in num_cols if c not in {"AttritionFlag"}]  # 排除标签列
    else:
        num_cols = numeric_override

    if categorical_override is None:
        cat_cols = df.select_dtypes(include=["object"]).columns.tolist()
    else:
        cat_cols = categorical_override
    cat_cols = [c for c in cat_cols if c not in {"Attrition", "AttritionFlag"}]  # 排除标签列

    # 数值特征管道：填充缺失值→标准化
    num_transformer = Pipeline([
        ("imputer", SimpleImputer(strategy="median")),
        ("scaler", StandardScaler())
    ])

    # 类别特征管道：填充缺失值→One-Hot编码
    cat_transformer = Pipeline([
        ("imputer", SimpleImputer(strategy="most_frequent")),
        ("ohe", onehot_encoder_compat())
    ])

    # 合并处理器
    preprocessor = ColumnTransformer(transformers=[
        ("num", num_transformer, num_cols),
        ("cat", cat_transformer, cat_cols)
    ], remainder="drop")

    logging.info(f"✅ 预处理器构建完成：数值特征{len(num_cols)}个，类别特征{len(cat_cols)}个")
    return preprocessor, num_cols, cat_cols


def get_preprocessor_feature_metadata(preprocessor):
    """提取预处理后特征与原始特征的映射关系"""
    transformed_feature_names = []
    raw_feature_names = []

    for name, trans, cols in preprocessor.transformers_:
        if name == "num":
            transformed_feature_names.extend(cols)
            raw_feature_names.extend(cols)
        elif name == "cat":
            ohe = trans.named_steps["ohe"]
            encoded_names = list(ohe.get_feature_names_out(cols))
            transformed_feature_names.extend(encoded_names)
            sorted_cols = sorted(cols, key=len, reverse=True)
            for encoded_name in encoded_names:
                mapped_col = next(
                    (col for col in sorted_cols if encoded_name == col or encoded_name.startswith(f"{col}_")),
                    encoded_name.split("_")[0]
                )
                raw_feature_names.append(mapped_col)

    return transformed_feature_names, raw_feature_names


def select_important_raw_features(df, fitted_preprocessor, fitted_lgb_model, min_keep_ratio=0.65, cumulative_threshold=0.92):
    """基于LightGBM重要性聚合到原始特征层面，筛选高价值特征"""
    _, raw_feature_names = get_preprocessor_feature_metadata(fitted_preprocessor)
    importances = np.asarray(fitted_lgb_model.feature_importances_, dtype=float)

    if len(importances) != len(raw_feature_names):
        logging.warning("⚠️ 特征重要性长度与预处理特征数不一致，跳过特征筛选")
        num_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        num_cols = [c for c in num_cols if c != "AttritionFlag"]
        cat_cols = df.select_dtypes(include=["object"]).columns.tolist()
        cat_cols = [c for c in cat_cols if c not in {"Attrition", "AttritionFlag"}]
        return num_cols, cat_cols, None

    importance_df = pd.DataFrame({
        "raw_feature": raw_feature_names,
        "importance": importances
    })
    feature_rank_df = (
        importance_df.groupby("raw_feature", as_index=False)["importance"]
        .sum()
        .sort_values("importance", ascending=False)
        .reset_index(drop=True)
    )

    positive_rank_df = feature_rank_df[feature_rank_df["importance"] > 0].copy()
    if positive_rank_df.empty:
        positive_rank_df = feature_rank_df.copy()

    total_importance = max(float(positive_rank_df["importance"].sum()), 1e-8)
    positive_rank_df["cum_ratio"] = positive_rank_df["importance"].cumsum() / total_importance

    min_keep = max(12, int(np.ceil(len(feature_rank_df) * min_keep_ratio)))
    min_keep = min(min_keep, len(feature_rank_df))

    selected_features = positive_rank_df[positive_rank_df["cum_ratio"] <= cumulative_threshold]["raw_feature"].tolist()
    if len(selected_features) < len(positive_rank_df):
        next_feature = positive_rank_df.iloc[len(selected_features)]["raw_feature"] if len(selected_features) < len(positive_rank_df) else None
        if next_feature and next_feature not in selected_features:
            selected_features.append(next_feature)

    if len(selected_features) < min_keep:
        selected_features = positive_rank_df.head(min_keep)["raw_feature"].tolist()

    selected_features = list(dict.fromkeys(selected_features))
    numeric_candidates = df.select_dtypes(include=[np.number]).columns.tolist()
    numeric_candidates = [c for c in numeric_candidates if c != "AttritionFlag"]
    categorical_candidates = df.select_dtypes(include=["object"]).columns.tolist()
    categorical_candidates = [c for c in categorical_candidates if c not in {"Attrition", "AttritionFlag"}]

    selected_num_cols = [col for col in numeric_candidates if col in selected_features]
    selected_cat_cols = [col for col in categorical_candidates if col in selected_features]

    if not selected_num_cols and not selected_cat_cols:
        logging.warning("⚠️ 特征筛选后为空，回退到原始特征集")
        selected_num_cols = numeric_candidates
        selected_cat_cols = categorical_candidates

    logging.info(
        "🔎 特征筛选完成：原始特征 %s 个 -> 保留 %s 个（数值 %s，类别 %s）",
        len(feature_rank_df), len(selected_num_cols) + len(selected_cat_cols),
        len(selected_num_cols), len(selected_cat_cols)
    )
    return selected_num_cols, selected_cat_cols, feature_rank_df


# -----------------------
# LightGBM自动调参（RandomizedSearch）
# -----------------------
def compute_scale_pos_weight(y):
    """根据标签分布计算正样本权重"""
    y_arr = np.asarray(y)
    pos_count = max(int(np.sum(y_arr == 1)), 1)
    neg_count = max(int(np.sum(y_arr == 0)), 1)
    return neg_count / pos_count


def evaluate_binary_probabilities(y_true, y_prob, threshold):
    """统一计算二分类概率输出在指定阈值下的关键指标"""
    threshold_array = np.asarray(threshold, dtype=float)
    if threshold_array.ndim == 0:
        threshold_array = np.full(len(y_prob), float(threshold_array), dtype=float)
    y_pred = (np.asarray(y_prob, dtype=float) >= threshold_array).astype(int)
    return {
        "acc": accuracy_score(y_true, y_pred),
        "precision": precision_score(y_true, y_pred, zero_division=0),
        "recall": recall_score(y_true, y_pred, zero_division=0),
        "f1": f1_score(y_true, y_pred, zero_division=0),
        "pred_positive_rate": float(np.mean(y_pred))
    }


def build_risk_segment_labels(df):
    """基于员工画像构建高风险/常规分层，用于分层阈值策略"""
    if df is None or len(df) == 0:
        return np.array([], dtype=object), np.array([], dtype=float)

    score = np.zeros(len(df), dtype=float)

    if 'OverTimeFlag' in df.columns:
        score += (df['OverTimeFlag'].fillna(0).to_numpy(dtype=float) >= 1).astype(float)
    if 'StressLoadScore' in df.columns:
        score += (df['StressLoadScore'].fillna(0).to_numpy(dtype=float) >= 2).astype(float)
    if 'SatisfactionIndex' in df.columns:
        score += (df['SatisfactionIndex'].fillna(df['SatisfactionIndex'].median()).to_numpy(dtype=float) <= 2.75).astype(float)
    if 'PromotionWaitRatio' in df.columns:
        score += (df['PromotionWaitRatio'].fillna(0).to_numpy(dtype=float) >= 0.45).astype(float)
    if 'RoleStagnationRatio' in df.columns:
        score += (df['RoleStagnationRatio'].fillna(0).to_numpy(dtype=float) >= 0.55).astype(float)
    if 'TravelRisk' in df.columns:
        score += (df['TravelRisk'].fillna(0).to_numpy(dtype=float) >= 1.5).astype(float)
    if 'macro_index' in df.columns:
        score += (df['macro_index'].fillna(df['macro_index'].median()).to_numpy(dtype=float) < 60).astype(float)
    if 'policy_net_support' in df.columns:
        score += (df['policy_net_support'].fillna(0).to_numpy(dtype=float) <= 0).astype(float)

    labels = np.where(score >= 2.0, "high_risk", "standard")
    high_share = float(np.mean(labels == "high_risk"))
    if high_share < 0.15 or high_share > 0.60:
        adaptive_cutoff = float(np.quantile(score, 0.65))
        labels = np.where(score >= max(adaptive_cutoff, 1.0), "high_risk", "standard")

    if len(np.unique(labels)) < 2:
        median_score = float(np.median(score))
        labels = np.where(score >= median_score, "high_risk", "standard")

    return labels, score


def resolve_threshold_array(y_prob, threshold_config, df_context=None):
    """将统一阈值或分层阈值配置展开成逐样本阈值数组"""
    if isinstance(threshold_config, dict) and threshold_config.get("type") == "segment":
        segment_labels, _ = build_risk_segment_labels(df_context)
        high_threshold = float(threshold_config["high_risk"])
        standard_threshold = float(threshold_config["standard"])
        threshold_array = np.where(segment_labels == "high_risk", high_threshold, standard_threshold).astype(float)
        return threshold_array, segment_labels

    base_threshold = float(threshold_config)
    return np.full(len(y_prob), base_threshold, dtype=float), None


def optimize_segment_thresholds(y_true, y_prob, df_context, base_threshold):
    """在高风险组/常规组上分别搜索阈值，优先提升整体F1"""
    segment_labels, segment_score = build_risk_segment_labels(df_context)
    if len(segment_labels) == 0:
        payload = evaluate_binary_probabilities(y_true, y_prob, base_threshold)
        return {
            "type": "global",
            "base_threshold": float(base_threshold),
            "global_threshold": float(base_threshold)
        }, payload

    high_mask = segment_labels == "high_risk"
    standard_mask = ~high_mask
    y_array = np.asarray(y_true)

    if (
        high_mask.sum() < max(30, int(0.12 * len(segment_labels)))
        or standard_mask.sum() < max(30, int(0.12 * len(segment_labels)))
        or int(np.sum(y_array[high_mask] == 1)) < 8
        or int(np.sum(y_array[standard_mask] == 1)) < 8
    ):
        payload = evaluate_binary_probabilities(y_true, y_prob, base_threshold)
        return {
            "type": "global",
            "base_threshold": float(base_threshold),
            "global_threshold": float(base_threshold)
        }, payload

    best_result = None
    high_candidates = np.arange(max(0.20, base_threshold - 0.18), min(0.75, base_threshold + 0.02) + 0.0001, 0.02)
    standard_candidates = np.arange(max(0.35, base_threshold - 0.02), min(0.90, base_threshold + 0.16) + 0.0001, 0.02)

    for high_threshold in high_candidates:
        for standard_threshold in standard_candidates:
            if high_threshold > standard_threshold:
                continue
            threshold_array = np.where(high_mask, high_threshold, standard_threshold).astype(float)
            payload = evaluate_binary_probabilities(y_true, y_prob, threshold_array)
            threshold_gap = standard_threshold - high_threshold
            score = (
                0.64 * payload["f1"]
                + 0.18 * payload["recall"]
                + 0.10 * payload["acc"]
                + 0.08 * payload["precision"]
                - 0.03 * abs(payload["pred_positive_rate"] - float(np.mean(y_true)))
                - 0.02 * max(threshold_gap - 0.18, 0.0)
            )
            if best_result is None or score > best_result["score"] + 1e-12:
                best_result = {
                    "threshold_config": {
                        "type": "segment",
                        "base_threshold": float(base_threshold),
                        "high_risk": float(high_threshold),
                        "standard": float(standard_threshold),
                        "high_risk_share": float(np.mean(high_mask)),
                        "high_risk_score_mean": float(np.mean(segment_score[high_mask])) if high_mask.any() else 0.0,
                        "standard_score_mean": float(np.mean(segment_score[standard_mask])) if standard_mask.any() else 0.0
                    },
                    "payload": payload,
                    "score": score
                }

    if best_result is None:
        payload = evaluate_binary_probabilities(y_true, y_prob, base_threshold)
        return {
            "type": "global",
            "base_threshold": float(base_threshold),
            "global_threshold": float(base_threshold)
        }, payload

    return best_result["threshold_config"], best_result["payload"]


def fit_probability_calibrator(y_true, y_prob):
    """基于OOF概率做一维Platt校准，提升阈值迁移稳定性"""
    clipped_prob = np.clip(np.asarray(y_prob, dtype=float), 1e-6, 1 - 1e-6)
    logit_feature = np.log(clipped_prob / (1.0 - clipped_prob)).reshape(-1, 1)
    calibrator = LogisticRegression(
        solver='lbfgs',
        C=1.0,
        random_state=RANDOM_STATE
    )
    calibrator.fit(logit_feature, np.asarray(y_true))
    return calibrator


def apply_probability_calibrator(calibrator, y_prob):
    """对概率应用Platt校准"""
    if calibrator is None:
        return np.asarray(y_prob, dtype=float)
    clipped_prob = np.clip(np.asarray(y_prob, dtype=float), 1e-6, 1 - 1e-6)
    logit_feature = np.log(clipped_prob / (1.0 - clipped_prob)).reshape(-1, 1)
    return calibrator.predict_proba(logit_feature)[:, 1]


def optimize_classification_threshold(y_true, y_prob):
    """在验证集上寻找更稳健的最佳阈值，避免过度保守压低召回"""
    actual_positive_rate = float(np.mean(y_true))
    target_positive_rate = float(np.clip(actual_positive_rate * 1.10, 0.12, 0.30))

    def evaluate_thresholds(thresholds, current_best_threshold=0.5, current_best_payload=None):
        best_threshold_local = current_best_threshold
        if current_best_payload is None:
            current_best_payload = {
                "score": -np.inf,
                "acc": 0.0,
                "precision": 0.0,
                "recall": 0.0,
                "f1": 0.0,
                "pred_positive_rate": 0.0
            }
        best_payload_local = current_best_payload

        for threshold in thresholds:
            payload = evaluate_binary_probabilities(y_true, y_prob, threshold)
            rate_gap = abs(payload["pred_positive_rate"] - target_positive_rate)
            score = (
                0.56 * payload["f1"]
                + 0.22 * payload["recall"]
                + 0.12 * payload["acc"]
                + 0.10 * payload["precision"]
                - 0.10 * rate_gap
                - 0.04 * max(float(threshold) - 0.62, 0.0)
            )
            if (
                score > best_payload_local["score"] + 1e-12
                or (
                    abs(score - best_payload_local["score"]) <= 1e-12
                    and (
                        payload["recall"] > best_payload_local["recall"] + 1e-12
                        or (
                            abs(payload["recall"] - best_payload_local["recall"]) <= 1e-12
                            and abs(threshold - 0.55) < abs(best_threshold_local - 0.55)
                        )
                    )
                )
            ):
                best_threshold_local = float(threshold)
                best_payload_local = dict(payload)
                best_payload_local["score"] = score

        return best_threshold_local, best_payload_local

    prevalence_threshold = float(np.quantile(y_prob, 1.0 - target_positive_rate))
    coarse_thresholds = np.unique(np.concatenate([
        np.arange(0.25, 0.701, 0.02),
        np.array([prevalence_threshold, 0.50, 0.55, 0.60])
    ]))
    best_threshold, best_payload = evaluate_thresholds(coarse_thresholds)
    fine_start = max(0.10, best_threshold - 0.05)
    fine_end = min(0.75, best_threshold + 0.05)
    fine_thresholds = np.arange(fine_start, fine_end + 0.0001, 0.005)
    best_threshold, best_payload = evaluate_thresholds(fine_thresholds, best_threshold, best_payload)
    return best_threshold, best_payload


def lgb_random_search(X_train_trans, y_train, n_iter=16, random_state=RANDOM_STATE):
    """LightGBM随机搜索调参"""
    ensure_lightgbm()
    scale_pos_weight = compute_scale_pos_weight(y_train)

    # 超参数搜索空间
    param_dist = {
        'num_leaves': [7, 15, 31],
        'learning_rate': [0.02, 0.03, 0.05],
        'n_estimators': [120, 180, 240, 320],
        'max_depth': [2, 3, 4],
        'min_child_samples': [35, 50, 70, 90],
        'subsample': [0.65, 0.75, 0.85],
        'colsample_bytree': [0.6, 0.7, 0.8],
        'reg_alpha': [0.5, 1.0, 2.0, 4.0],
        'reg_lambda': [1.0, 2.0, 4.0, 6.0],
        'min_split_gain': [0.0, 0.1, 0.2, 0.4]
    }

    # 初始化模型与搜索
    clf = lgb.LGBMClassifier(
        objective='binary',
        random_state=random_state,
        n_jobs=1,
        scale_pos_weight=scale_pos_weight
    )
    cv = StratifiedKFold(n_splits=4, shuffle=True, random_state=random_state)
    rs = RandomizedSearchCV(
        clf,
        param_distributions=param_dist,
        n_iter=n_iter,
        scoring='roc_auc',
        cv=cv,
        n_jobs=-1,
        random_state=random_state,
        verbose=0
    )
    rs.fit(X_train_trans, y_train)

    logging.info(f"✅ LightGBM调参完成，最佳参数：{rs.best_params_}")
    return rs.best_estimator_


def build_base_models(best_lgb_params, scale_pos_weight):
    """构建基础模型：更保守的LightGBM + 平衡LR + ExtraTrees"""
    lgb_param_keys = [
        "num_leaves", "learning_rate", "n_estimators", "max_depth",
        "min_child_samples", "subsample", "colsample_bytree",
        "reg_alpha", "reg_lambda", "min_split_gain"
    ]
    lgb_params = {
        "objective": "binary",
        "random_state": RANDOM_STATE,
        "n_jobs": -1,
        "scale_pos_weight": scale_pos_weight
    }
    for key in lgb_param_keys:
        if key in best_lgb_params:
            lgb_params[key] = best_lgb_params[key]

    lgb_clf = lgb.LGBMClassifier(**lgb_params)
    lr_clf = LogisticRegression(
        max_iter=1000,
        solver='liblinear',
        class_weight='balanced',
        C=0.35,
        random_state=RANDOM_STATE
    )
    et_clf = ExtraTreesClassifier(
        n_estimators=400,
        max_depth=7,
        min_samples_split=12,
        min_samples_leaf=5,
        max_features='sqrt',
        class_weight='balanced_subsample',
        random_state=RANDOM_STATE,
        n_jobs=-1
    )
    return {"lgb": lgb_clf, "lr": lr_clf, "et": et_clf}


def build_meta_feature_matrix(prob_map, weight_map):
    """基于基础模型概率构造二层融合特征"""
    lgb_prob = np.asarray(prob_map["lgb"], dtype=float)
    lr_prob = np.asarray(prob_map["lr"], dtype=float)
    et_prob = np.asarray(prob_map["et"], dtype=float)
    base_matrix = np.column_stack([lgb_prob, lr_prob, et_prob])
    blended_prob = (
        weight_map["lgb"] * lgb_prob
        + weight_map["lr"] * lr_prob
        + weight_map["et"] * et_prob
    )
    mean_prob = base_matrix.mean(axis=1)
    std_prob = base_matrix.std(axis=1)
    disagreement = base_matrix.max(axis=1) - base_matrix.min(axis=1)
    return np.column_stack([
        lgb_prob,
        lr_prob,
        et_prob,
        blended_prob,
        mean_prob,
        std_prob,
        disagreement
    ])


def build_meta_learner():
    """构建轻量二层融合器，避免固定权重在测试集迁移失真"""
    return Pipeline([
        ("scaler", StandardScaler()),
        ("lr", LogisticRegression(
            max_iter=1000,
            solver='lbfgs',
            C=0.35,
            class_weight='balanced',
            random_state=RANDOM_STATE
        ))
    ])


def fit_meta_learner_with_oof(meta_X, y, n_splits=5):
    """基于OOF二层特征训练轻量元模型，并返回元模型OOF预测"""
    y_array = np.asarray(y)
    meta_oof_prob = np.zeros(len(y_array), dtype=float)
    cv = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=RANDOM_STATE)

    for train_idx, valid_idx in cv.split(meta_X, y_array):
        meta_model_fold = build_meta_learner()
        meta_model_fold.fit(meta_X[train_idx], y_array[train_idx])
        meta_oof_prob[valid_idx] = meta_model_fold.predict_proba(meta_X[valid_idx])[:, 1]

    final_meta_model = build_meta_learner()
    final_meta_model.fit(meta_X, y_array)
    return meta_oof_prob, final_meta_model


class BlendedAttritionModel:
    """基础模型融合器，可选轻量元模型做二层决策"""
    def __init__(self, model_map, weight_map, calibrator=None, meta_model=None):
        self.named_estimators_ = model_map
        weight_sum = sum(weight_map.values()) if weight_map else 1.0
        self.blend_weights = {
            name: float(weight / weight_sum)
            for name, weight in weight_map.items()
        }
        self.calibrator = calibrator
        self.meta_model = meta_model

    def predict_proba(self, X):
        prob_map = {}
        for name, model in self.named_estimators_.items():
            prob_map[name] = model.predict_proba(X)[:, 1]

        blended_prob = (
            self.blend_weights.get("lgb", 0.0) * prob_map["lgb"]
            + self.blend_weights.get("lr", 0.0) * prob_map["lr"]
            + self.blend_weights.get("et", 0.0) * prob_map["et"]
        )
        if self.meta_model is not None:
            meta_X = build_meta_feature_matrix(prob_map, self.blend_weights)
            blended_prob = self.meta_model.predict_proba(meta_X)[:, 1]
        blended_prob = apply_probability_calibrator(self.calibrator, blended_prob)
        return np.column_stack([1 - blended_prob, blended_prob])

    def predict(self, X, threshold=0.5):
        return (self.predict_proba(X)[:, 1] >= threshold).astype(int)


def generate_oof_predictions(X, y, best_lgb_params, n_splits=5):
    """生成多个基础模型的OOF预测，用于稳健融合与阈值优化"""
    y_array = np.asarray(y)
    oof_pred_map = {
        "lgb": np.zeros(len(y_array)),
        "lr": np.zeros(len(y_array)),
        "et": np.zeros(len(y_array))
    }
    cv = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=RANDOM_STATE)

    for train_idx, valid_idx in cv.split(X, y_array):
        X_train_fold, X_valid_fold = X[train_idx], X[valid_idx]
        y_train_fold = y_array[train_idx]

        model_map = build_base_models(best_lgb_params, compute_scale_pos_weight(y_train_fold))
        for name, model in model_map.items():
            model.fit(X_train_fold, y_train_fold)
            oof_pred_map[name][valid_idx] = model.predict_proba(X_valid_fold)[:, 1]

    return oof_pred_map


def optimize_blend_and_threshold(y_true, prob_map):
    """联合寻找最佳融合权重和分类阈值"""
    def search_weight_grid(step, centers=None, current_best=None):
        if current_best is None:
            current_best = {
                "weights": {"lgb": 0.5, "lr": 0.3, "et": 0.2},
                "threshold": 0.5,
                "metrics": {"auc": 0.0, "acc": 0.0, "f1": 0.0},
                "score": -np.inf
            }

        if centers is None:
            lgb_values = np.arange(0.0, 1.0001, step)
            lr_values = np.arange(0.0, 1.0001, step)
        else:
            lgb_values = np.arange(max(0.0, centers["lgb"] - step * 2), min(1.0, centers["lgb"] + step * 2) + 0.0001, step)
            lr_values = np.arange(max(0.0, centers["lr"] - step * 2), min(1.0, centers["lr"] + step * 2) + 0.0001, step)

        for w_lgb in lgb_values:
            for w_lr in lr_values:
                w_et = round(1.0 - w_lgb - w_lr, 10)
                if w_et < 0 or w_et > 1:
                    continue
                weight_map = {"lgb": float(w_lgb), "lr": float(w_lr), "et": float(w_et)}
                if max(weight_map.values()) < 0.35:
                    continue

                blended_prob = (
                    weight_map["lgb"] * prob_map["lgb"]
                    + weight_map["lr"] * prob_map["lr"]
                    + weight_map["et"] * prob_map["et"]
                )
                auc = roc_auc_score(y_true, blended_prob)
                threshold, payload = optimize_classification_threshold(y_true, blended_prob)
                rate_gap = abs(payload["pred_positive_rate"] - float(np.mean(y_true)))
                complexity_penalty = (
                    0.05 * max(weight_map["lgb"] - 0.35, 0.0)
                    + 0.04 * max(weight_map["et"] - 0.20, 0.0)
                    + 0.04 * max(0.45 - weight_map["lr"], 0.0)
                )
                score = (
                    0.30 * auc
                    + 0.38 * payload["f1"]
                    + 0.20 * payload["recall"]
                    + 0.07 * payload["acc"]
                    + 0.05 * payload["precision"]
                    - 0.07 * rate_gap
                    - complexity_penalty
                )
                if score > current_best["score"] + 1e-12:
                    current_best = {
                        "weights": weight_map,
                        "threshold": float(threshold),
                        "metrics": {
                            "auc": auc,
                            "acc": payload["acc"],
                            "precision": payload["precision"],
                            "recall": payload["recall"],
                            "f1": payload["f1"],
                            "pred_positive_rate": payload["pred_positive_rate"]
                        },
                        "score": score
                    }
        return current_best

    best_result = search_weight_grid(0.05)
    best_result = search_weight_grid(0.02, centers=best_result["weights"], current_best=best_result)
    return best_result["weights"], best_result["threshold"], best_result["metrics"]


# -----------------------
# 融合模型训练（LightGBM+逻辑回归+ExtraTrees）
# -----------------------
def train_stacking_lgb(X_df, y, preprocessor):
    """训练调参与OOF融合优化后的集成模型（使用全量增强特征）"""
    ensure_lightgbm()
    # 划分训练池与测试集
    X_train_valid_df, X_test_df, y_train_valid, y_test = train_test_split(
        X_df, y, test_size=TEST_SIZE, stratify=y, random_state=RANDOM_STATE
    )
    y_train_valid_array = np.asarray(y_train_valid)

    # 1. 使用全量增强特征，避免激进筛选带来的泛化损失
    selected_num_cols = X_train_valid_df.select_dtypes(include=[np.number]).columns.tolist()
    selected_num_cols = [c for c in selected_num_cols if c != "AttritionFlag"]
    selected_cat_cols = X_train_valid_df.select_dtypes(include=["object"]).columns.tolist()
    selected_cat_cols = [c for c in selected_cat_cols if c not in {"Attrition", "AttritionFlag"}]
    preprocessor = clone(preprocessor)
    X_train_valid_trans = preprocessor.fit_transform(X_train_valid_df)
    X_test_trans = preprocessor.transform(X_test_df)

    # 2. 在增强后的全量特征空间中调参与融合
    best_lgb = lgb_random_search(X_train_valid_trans, y_train_valid_array)
    best_lgb_params = best_lgb.get_params()

    # 3. 生成OOF预测，寻找最佳融合权重和阈值
    oof_pred_map = generate_oof_predictions(X_train_valid_trans, y_train_valid_array, best_lgb_params)
    best_weight_map, _, _ = optimize_blend_and_threshold(y_train_valid_array, oof_pred_map)
    meta_oof_X = build_meta_feature_matrix(oof_pred_map, best_weight_map)
    meta_oof_prob, meta_model = fit_meta_learner_with_oof(meta_oof_X, y_train_valid_array)
    best_threshold, calibrated_oof_metrics = optimize_classification_threshold(y_train_valid_array, meta_oof_prob)
    threshold_strategy, segmented_oof_metrics = optimize_segment_thresholds(
        y_train_valid_array, meta_oof_prob, X_train_valid_df, best_threshold
    )
    oof_metrics = {
        "auc": roc_auc_score(y_train_valid_array, meta_oof_prob),
        "acc": segmented_oof_metrics["acc"],
        "precision": segmented_oof_metrics["precision"],
        "recall": segmented_oof_metrics["recall"],
        "f1": segmented_oof_metrics["f1"],
        "pred_positive_rate": segmented_oof_metrics["pred_positive_rate"]
    }

    # 4. 用全部训练池重训最终基础模型
    final_model_map = build_base_models(best_lgb_params, compute_scale_pos_weight(y_train_valid_array))
    for model in final_model_map.values():
        model.fit(X_train_valid_trans, y_train_valid_array)
    blended_model = BlendedAttritionModel(final_model_map, best_weight_map, calibrator=None, meta_model=meta_model)

    # 5. 预测概率
    y_train_prob = blended_model.predict_proba(X_train_valid_trans)[:, 1]
    y_test_prob = blended_model.predict_proba(X_test_trans)[:, 1]
    train_thresholds, train_segment_labels = resolve_threshold_array(y_train_prob, threshold_strategy, X_train_valid_df)
    test_thresholds, test_segment_labels = resolve_threshold_array(y_test_prob, threshold_strategy, X_test_df)
    train_eval = evaluate_binary_probabilities(y_train_valid_array, y_train_prob, train_thresholds)
    test_eval = evaluate_binary_probabilities(np.asarray(y_test), y_test_prob, test_thresholds)

    # 6. 计算评估指标
    metrics = {
        'train_auc': roc_auc_score(y_train_valid, y_train_prob),
        'valid_auc': oof_metrics['auc'],
        'test_auc': roc_auc_score(y_test, y_test_prob),
        'train_acc': train_eval['acc'],
        'valid_acc': oof_metrics['acc'],
        'test_acc': test_eval['acc'],
        'train_precision': train_eval['precision'],
        'valid_precision': oof_metrics['precision'],
        'test_precision': test_eval['precision'],
        'train_recall': train_eval['recall'],
        'valid_recall': oof_metrics['recall'],
        'test_recall': test_eval['recall'],
        'train_pred_positive_rate': train_eval['pred_positive_rate'],
        'valid_pred_positive_rate': oof_metrics['pred_positive_rate'],
        'test_pred_positive_rate': test_eval['pred_positive_rate'],
        'train_f1': train_eval['f1'],
        'valid_f1': oof_metrics['f1'],
        'test_f1': test_eval['f1'],
        'blend_lgb_weight': best_weight_map['lgb'],
        'blend_lr_weight': best_weight_map['lr'],
        'blend_et_weight': best_weight_map['et'],
        'ensemble_strategy': 'oof_logistic_stack',
        'feature_strategy': 'full_enhanced_features',
        'probability_calibration': 'none',
        'threshold_strategy': threshold_strategy.get('type', 'global'),
        'best_threshold': float(threshold_strategy.get('base_threshold', best_threshold)),
        'high_risk_threshold': float(threshold_strategy.get('high_risk', threshold_strategy.get('global_threshold', best_threshold))),
        'standard_threshold': float(threshold_strategy.get('standard', threshold_strategy.get('global_threshold', best_threshold))),
        'train_high_risk_share': float(np.mean(train_segment_labels == "high_risk")) if train_segment_labels is not None and len(train_segment_labels) else 0.0,
        'test_high_risk_share': float(np.mean(test_segment_labels == "high_risk")) if test_segment_labels is not None and len(test_segment_labels) else 0.0,
        'selected_num_features': len(selected_num_cols),
        'selected_cat_features': len(selected_cat_cols),
        'selected_total_features': len(selected_num_cols) + len(selected_cat_cols),
    }

    logging.info("✅ 模型训练完成，评估结果：")
    logging.info(
        "  - 使用增强原始特征：数值 %s 个 | 类别 %s 个 | 合计 %s 个",
        metrics['selected_num_features'], metrics['selected_cat_features'], metrics['selected_total_features']
    )
    logging.info(
        "  - OOF-AUC：%.4f | OOF-Acc：%.4f | OOF-F1：%.4f | 融合权重(LGB/LR/ET)=%.2f/%.2f/%.2f | 最优阈值：%.2f",
        metrics['valid_auc'], metrics['valid_acc'], metrics['valid_f1'],
        best_weight_map['lgb'], best_weight_map['lr'], best_weight_map['et'], metrics['best_threshold']
    )
    logging.info("  - 集成策略：%s", metrics['ensemble_strategy'])
    logging.info("  - 概率校准方式：%s", metrics['probability_calibration'])
    logging.info(
        "  - 阈值策略：%s | 高风险阈值：%.2f | 常规阈值：%.2f | 测试高风险组占比：%.4f",
        metrics['threshold_strategy'], metrics['high_risk_threshold'], metrics['standard_threshold'], metrics['test_high_risk_share']
    )
    logging.info(
        "  - OOF-Precision：%.4f | OOF-Recall：%.4f | OOF预测流失率：%.4f",
        metrics['valid_precision'], metrics['valid_recall'], metrics['valid_pred_positive_rate']
    )
    logging.info("  - 训练AUC：%.4f | 测试AUC：%.4f", metrics['train_auc'], metrics['test_auc'])
    logging.info("  - 训练Acc：%.4f | 测试Acc：%.4f", metrics['train_acc'], metrics['test_acc'])
    logging.info(
        "  - 训练Precision：%.4f | 测试Precision：%.4f",
        metrics['train_precision'], metrics['test_precision']
    )
    logging.info(
        "  - 训练Recall：%.4f | 测试Recall：%.4f | 测试预测流失率：%.4f",
        metrics['train_recall'], metrics['test_recall'], metrics['test_pred_positive_rate']
    )
    logging.info("  - 训练F1：%.4f | 测试F1：%.4f", metrics['train_f1'], metrics['test_f1'])

    return blended_model, preprocessor, X_train_valid_df, X_test_df, y_train_valid, y_test, metrics, threshold_strategy


# -----------------------
# SHAP分析（Top3风险驱动+可视化）
# -----------------------
def compute_shap_top3_and_export(model, preprocessor, X_test_df, df_test, out_prefix="employee_risk"):
    """计算每个员工Top3风险驱动特征，并保存SHAP图"""
    if shap is None:
        logging.warning("❌ shap未安装，跳过SHAP分析")
        return None

    # 添加调试信息
    logging.info("🔍 SHAP分析开始...")

    # 输出路径（当前目录）
    out_prefix = os.path.join(CURRENT_DIR, out_prefix)
    os.makedirs(os.path.dirname(out_prefix), exist_ok=True)

    # 预处理测试数据
    X_test_trans = preprocessor.transform(X_test_df)
    logging.info(f"🔍 测试数据形状: {X_test_trans.shape}")

    # 获取特征名
    feature_names = []
    try:
        num_cols = preprocessor.transformers_[0][2]
        cat_transformer = preprocessor.transformers_[1][1]
        cat_cols = preprocessor.transformers_[1][2]
        ohe = cat_transformer.named_steps["ohe"]
        feature_names = list(num_cols) + list(ohe.get_feature_names_out(cat_cols))
        logging.info(f"🔍 特征数量: {len(feature_names)}")
    except Exception as e:
        logging.warning(f"⚠️ 无法自动获取特征名: {e}")
        feature_names = [f"特征_{i}" for i in range(X_test_trans.shape[1])]

    # 计算SHAP值 - 添加详细调试
    try:
        logging.info("🔍 尝试获取LightGBM子模型...")
        # 调试：打印模型结构
        logging.info(f"🔍 模型类型: {type(model)}")
        logging.info(f"🔍 模型命名估计器: {list(model.named_estimators_.keys())}")

        lgb_model = model.named_estimators_['lgb']
        logging.info(f"🔍 LightGBM模型类型: {type(lgb_model)}")

        logging.info("🔍 创建SHAP解释器...")
        explainer = shap.TreeExplainer(lgb_model)

        logging.info("🔍 计算SHAP值...")
        shap_vals = explainer.shap_values(X_test_trans)
        logging.info(f"🔍 SHAP值类型: {type(shap_vals)}")

        # 处理SHAP值格式
        if isinstance(shap_vals, list):
            shap_arr = shap_vals[1]  # 二分类取正类SHAP值
            logging.info("🔍 使用列表格式的SHAP值(二分类)")
        else:
            shap_arr = shap_vals
            logging.info("🔍 使用数组格式的SHAP值")

        logging.info(f"🔍 SHAP数组形状: {shap_arr.shape}")

    except Exception as e:
        logging.error(f"❌ SHAP值计算失败：{e}")
        import traceback
        logging.error(traceback.format_exc())  # 打印完整堆栈跟踪
        return None

    # ========== 以下是缺失的关键部分 ==========

    # 计算每个样本的Top3风险驱动特征
    top3_list = []
    for i in range(shap_arr.shape[0]):
        row = shap_arr[i]
        # 按SHAP绝对值排序，取前3
        idxs = np.argsort(np.abs(row))[::-1][:3]
        items = [(feature_names[j] if j < len(feature_names) else f"特征_{j}", float(row[j])) for j in idxs]
        top3_list.append(items)

    # 格式化Top3特征
    def fmt_top3(x):
        return "; ".join([f"{t[0]}({t[1]:.3f})" for t in x])

    df_test["Top3_风险驱动"] = [fmt_top3(x) for x in top3_list]

    # 保存SHAP summary图
    try:
        plt.figure(figsize=(10, 6))
        shap.summary_plot(shap_arr, X_test_trans, feature_names=feature_names, show=False)
        shap_save_path = f"{out_prefix}_shap_summary.png"
        plt.tight_layout()
        plt.savefig(shap_save_path, dpi=150, bbox_inches='tight')
        plt.close()
        logging.info(f"✅ SHAP图已保存：{shap_save_path}")
    except Exception as e:
        logging.warning(f"❌ SHAP绘图失败：{e}")

    return df_test


METRIC_LABEL_MAP = {
    "train_auc": "训练集AUC",
    "valid_auc": "验证集AUC",
    "test_auc": "测试集AUC",
    "train_acc": "训练集Accuracy",
    "valid_acc": "验证集Accuracy",
    "test_acc": "测试集Accuracy",
    "train_f1": "训练集F1",
    "valid_f1": "验证集F1",
    "test_f1": "测试集F1",
    "train_precision": "训练集Precision",
    "valid_precision": "验证集Precision",
    "test_precision": "测试集Precision",
    "train_recall": "训练集Recall",
    "valid_recall": "验证集Recall",
    "test_recall": "测试集Recall",
    "train_pred_positive_rate": "训练集预测流失率",
    "valid_pred_positive_rate": "验证集预测流失率",
    "test_pred_positive_rate": "测试集预测流失率",
    "best_threshold": "基础阈值",
    "high_risk_threshold": "高风险组阈值",
    "standard_threshold": "常规组阈值",
    "train_high_risk_share": "训练集高风险占比",
    "test_high_risk_share": "测试集高风险占比",
    "blend_lgb_weight": "融合权重-LGB",
    "blend_lr_weight": "融合权重-LR",
    "blend_et_weight": "融合权重-ET",
    "threshold_strategy": "阈值策略",
    "ensemble_strategy": "集成策略",
    "feature_strategy": "特征策略",
    "probability_calibration": "概率校准",
    "selected_num_features": "数值特征数量",
    "selected_cat_features": "类别特征数量",
    "selected_total_features": "总特征数量",
}


def build_metrics_export_frames(metrics):
    """构建更易读的指标导出表（核心指标 + 完整指标）。"""
    core_keys = [
        "train_auc", "valid_auc", "test_auc",
        "train_acc", "valid_acc", "test_acc",
        "train_f1", "valid_f1", "test_f1",
        "train_precision", "valid_precision", "test_precision",
        "train_recall", "valid_recall", "test_recall",
        "test_pred_positive_rate",
        "best_threshold", "high_risk_threshold", "standard_threshold",
        "train_high_risk_share", "test_high_risk_share",
    ]

    core_rows = []
    for key in core_keys:
        if key not in metrics:
            continue
        core_rows.append({
            "指标代码": key,
            "指标名称": METRIC_LABEL_MAP.get(key, key),
            "指标值": metrics[key],
        })

    full_rows = []
    for key, value in metrics.items():
        full_rows.append({
            "指标代码": key,
            "指标名称": METRIC_LABEL_MAP.get(key, key),
            "指标值": value,
        })

    return pd.DataFrame(core_rows), pd.DataFrame(full_rows)

# -----------------------
# 最终结果输出（预测名单+政策缺口）
# -----------------------
def generate_outputs_and_reports(df_emp, model, preprocessor, X_test_df, y_test, metrics, threshold=0.5, out_prefix="employee_risk"):
    """输出员工风险预测名单、Top3风险驱动、政策缺口岗位"""
    # 输出路径（当前目录）
    out_prefix = os.path.join(CURRENT_DIR, out_prefix)

    # 1. 预测测试集风险
    X_test_trans = preprocessor.transform(X_test_df)
    y_pred_prob = model.predict_proba(X_test_trans)[:, 1]
    threshold_array, segment_labels = resolve_threshold_array(y_pred_prob, threshold, X_test_df)
    y_pred_label = (y_pred_prob >= threshold_array).astype(int)

    # 整理输出数据
    df_out = X_test_df.copy()
    df_out['流失概率'] = y_pred_prob.round(3)
    df_out['预测流失标签'] = y_pred_label
    df_out['实际流失标签'] = y_test.values
    df_out['预测阈值'] = np.round(threshold_array, 3)
    if segment_labels is not None:
        df_out['风险分层'] = segment_labels

    df_out_sorted = df_out.sort_values("流失概率", ascending=False).reset_index(drop=True)
    high_risk_df = df_out_sorted[df_out_sorted["预测流失标签"] == 1].copy()
    if high_risk_df.empty:
        high_risk_df = df_out_sorted.head(min(30, len(df_out_sorted))).copy()

    threshold_desc = threshold.get("type", "global") if isinstance(threshold, dict) else "global"
    summary_df = pd.DataFrame([
        {"指标": "测试样本数", "值": int(len(df_out_sorted)), "说明": "本次用于预测的员工样本数"},
        {"指标": "高风险人数(预测)", "值": int(df_out_sorted["预测流失标签"].sum()), "说明": "预测流失标签=1的人数"},
        {"指标": "高风险占比(预测)", "值": float(np.mean(df_out_sorted["预测流失标签"])), "说明": "高风险人数 / 测试样本数"},
        {"指标": "平均流失概率", "值": float(df_out_sorted["流失概率"].mean()), "说明": "测试集员工平均流失概率"},
        {"指标": "阈值策略", "值": threshold_desc, "说明": "模型使用的风险阈值方案"},
    ])

    # 保存预测名单（友好版）
    pred_save_path = f"{out_prefix}_预测结果.xlsx"
    save_friendly_excel(
        pred_save_path,
        sheet_frames={
            "预测明细": df_out_sorted,
            "高风险名单": high_risk_df,
            "结果摘要": summary_df,
        },
        percent_cols_map={
            "预测明细": ["流失概率", "预测阈值"],
            "高风险名单": ["流失概率", "预测阈值"],
            "结果摘要": ["值"],
        },
        heatmap_cols_map={
            "预测明细": ["流失概率"],
            "高风险名单": ["流失概率"],
        },
    )
    logging.info(f"✅ 员工风险预测名单已保存：{pred_save_path}")

    # 1.1 去留预测可视化图
    decision_view_name = f"{os.path.basename(out_prefix)}_去留预测可视化.png"
    plot_attrition_decision_view(
        df_out_sorted,
        threshold=threshold,
        save_name=decision_view_name,
        alias_names=["attrition_decision_view.png"],
    )

    # 2. SHAP Top3风险驱动
    if shap is not None:
        df_out_shap = compute_shap_top3_and_export(model, preprocessor, X_test_df, df_out_sorted.copy(), out_prefix)
        if df_out_shap is not None:
            shap_save_path = f"{out_prefix}_Top3风险驱动.xlsx"
            df_out_shap = df_out_shap.sort_values("流失概率", ascending=False).reset_index(drop=True)
            save_friendly_excel(
                shap_save_path,
                sheet_frames={
                    "Top3驱动明细": df_out_shap,
                },
                percent_cols_map={
                    "Top3驱动明细": ["流失概率", "预测阈值"],
                },
                heatmap_cols_map={
                    "Top3驱动明细": ["流失概率"],
                },
            )
            logging.info(f"✅ Top3风险驱动名单已保存：{shap_save_path}")

    # 3. 政策缺口岗位（macro_index < 50）
    if 'macro_index' in df_emp.columns:
        policy_gap = df_emp[df_emp['macro_index'] < 50][['JobRole', 'Department', 'macro_index']].drop_duplicates()
        if not policy_gap.empty:
            gap_save_path = f"{out_prefix}_政策缺口岗位.xlsx"
            policy_gap = policy_gap.sort_values("macro_index", ascending=True).reset_index(drop=True)
            save_friendly_excel(
                gap_save_path,
                sheet_frames={
                    "政策缺口岗位": policy_gap,
                }
            )
            logging.info(f"✅ 政策缺口岗位名单已保存：{gap_save_path}")
        else:
            logging.info("ℹ️  无政策缺口岗位（所有岗位macro_index ≥ 50）")

    # 4. 保存模型评估指标（友好版）
    metrics_save_path = f"{out_prefix}_模型评估指标.xlsx"
    core_metrics_df, full_metrics_df = build_metrics_export_frames(metrics)
    save_friendly_excel(
        metrics_save_path,
        sheet_frames={
            "核心指标": core_metrics_df,
            "完整指标": full_metrics_df,
        },
        percent_cols_map={
            "核心指标": ["指标值"],
            "完整指标": ["指标值"],
        },
    )
    logging.info(f"✅ 模型评估指标已保存：{metrics_save_path}")

# -----------------------
# 导出Top20特征重要性（表格）
# -----------------------
def export_top20_features(model, preprocessor, save_name="feature_importance_top20.xlsx"):
    """导出Top20特征重要性表格"""
    save_path = os.path.join(CURRENT_DIR, save_name)
    try:
        lgb_model = model.named_estimators_['lgb']
        # 获取特征名
        feature_names = []
        for name, trans, cols in preprocessor.transformers_:
            if name == 'num':
                feature_names.extend(cols)
            elif name == 'cat':
                ohe = trans.named_steps['ohe']
                feature_names.extend(ohe.get_feature_names_out(cols))
        # 计算重要性并排序
        importances = lgb_model.feature_importances_
        df_feat = pd.DataFrame({'特征名称': feature_names, '重要性得分': importances})
        df_feat = df_feat.sort_values('重要性得分', ascending=False).head(20)
        # 保存表格（友好版）
        save_friendly_excel(
            save_path,
            sheet_frames={"Top20特征重要性": df_feat}
        )
        logging.info(f"✅ Top20特征重要性表格已保存：{save_path}")
    except Exception as e:
        logging.warning(f"❌ 导出Top20特征失败：{e}")


def collect_generated_files(out_prefix="employee_attrition_analysis"):
    """收集当前运行目录内的核心输出文件路径。"""
    candidates = [
        os.path.join(CURRENT_DIR, f"{out_prefix}_预测结果.xlsx"),
        os.path.join(CURRENT_DIR, f"{out_prefix}_Top3风险驱动.xlsx"),
        os.path.join(CURRENT_DIR, f"{out_prefix}_政策缺口岗位.xlsx"),
        os.path.join(CURRENT_DIR, f"{out_prefix}_模型评估指标.xlsx"),
        os.path.join(CURRENT_DIR, f"{out_prefix}_shap_summary.png"),
        os.path.join(CURRENT_DIR, f"{out_prefix}_去留预测可视化.png"),
        os.path.join(CURRENT_DIR, "feature_importance_top20.xlsx"),
        os.path.join(CURRENT_DIR, "model_metrics.png"),
        os.path.join(CURRENT_DIR, "model-metric.png"),
        os.path.join(CURRENT_DIR, "attrition_decision_view.png"),
        os.path.join(CURRENT_DIR, "feature_importance.png"),
        os.path.join(CURRENT_DIR, "attrition_risk_distribution.png"),
        os.path.join(CURRENT_DIR, "policy_job_matching.png"),
    ]

    matched_by_prefix = sorted(glob.glob(os.path.join(CURRENT_DIR, f"{out_prefix}*")))
    merged = candidates + matched_by_prefix

    results = []
    for path in merged:
        abs_path = os.path.abspath(path)
        if os.path.exists(abs_path) and abs_path not in results:
            results.append(abs_path)
    return results


def run_pipeline(employee_data_path=None, policy_data_path=None, output_dir=None, out_prefix="employee_attrition_analysis"):
    """可复用总入口：支持脚本/网页统一调用。"""
    employee_path = os.path.abspath(employee_data_path or DATA_PATH)
    policy_path = os.path.abspath(policy_data_path or POLICY_PATH)
    runtime_dir = os.path.abspath(output_dir or CURRENT_DIR)

    try:
        configure_runtime_paths(current_dir=runtime_dir, data_path=employee_path, policy_path=policy_path)
        ensure_lightgbm()
        ensure_shap_warn()

        logging.info("=" * 50)
        logging.info("🎯 员工流失预测与政策匹配分析系统启动")
        logging.info("=" * 50)
        logging.info("运行目录：%s", CURRENT_DIR)
        logging.info("员工数据：%s", DATA_PATH)
        logging.info("政策数据：%s", POLICY_PATH)

        # 1. 初始化BERT模型（语义编码用）
        logging.info("\n1. 加载BERT语义模型...")
        tokenizer, bert_model = load_bert_model()

        # 2. 加载并预处理员工数据
        logging.info("\n2. 加载并预处理员工数据...")
        df_emp = load_and_preprocess_employee(DATA_PATH)
        df_emp = add_interaction_features(df_emp)  # 添加交互特征

        # 3. 加载政策数据并添加政策语义特征
        logging.info("\n3. 处理政策数据并生成语义特征...")
        policy_df = prepare_policy_dataframe(POLICY_PATH, tokenizer, bert_model)
        df_emp = add_policy_effect(df_emp, policy_df, tokenizer, bert_model)

        # 4. 构建宏观政策指数
        logging.info("\n4. 构建宏观政策指数...")
        policy_grouped = build_policy_macro_index_enhanced(policy_df, tokenizer, bert_model)
        if not policy_grouped.empty:
            if "JobRoleKey" in policy_grouped.columns and "JobRoleKey" in df_emp.columns:
                df_emp = df_emp.merge(policy_grouped[["JobRoleKey", "macro_index"]], on="JobRoleKey", how="left")
                df_emp["macro_index"] = df_emp["macro_index"].fillna(policy_grouped["macro_index"].mean())
            else:
                df_emp["macro_index"] = policy_grouped["macro_index"].iloc[0]
        else:
            df_emp["macro_index"] = 50.0  # 默认值
        df_emp.drop(columns=["JobRoleKey", "DepartmentKey"], inplace=True, errors="ignore")

        # 5. 构建数据预处理器
        logging.info("\n5. 构建数据预处理器...")
        preprocessor, num_cols, cat_cols = build_preprocessor(df_emp)

        # 6. 训练Stacking模型
        logging.info("\n6. 训练Stacking集成模型...")
        model, preprocessor, X_train_df, X_test_df, y_train, y_test, metrics, best_threshold = train_stacking_lgb(
            df_emp.drop(columns=["AttritionFlag"]),  # 特征集
            df_emp["AttritionFlag"],  # 标签集
            preprocessor
        )

        # 7. 生成所有输出文件
        logging.info("\n7. 生成结果报告与文件...")
        generate_outputs_and_reports(
            df_emp, model, preprocessor, X_test_df, y_test, metrics,
            threshold=best_threshold, out_prefix=out_prefix
        )
        export_top20_features(model, preprocessor)

        # 8. 生成可视化图表
        logging.info("\n8. 生成可视化图表...")
        # 模型性能图
        plot_model_metrics(metrics)
        # 特征重要性图
        plot_feature_importance(model, preprocessor)
        # 风险分布直方图
        y_pred_prob = model.predict_proba(preprocessor.transform(X_test_df))[:, 1]
        plot_attrition_risk_distribution(y_pred_prob, threshold=best_threshold)
        # 政策-岗位匹配图
        total_policy_score, policy_post_mapping, _ = compute_policy_impact(policy_df, tokenizer, bert_model)
        plot_policy_job_matching(policy_post_mapping)

        output_files = collect_generated_files(out_prefix=out_prefix)

        logging.info("\n" + "=" * 50)
        logging.info("🎉 所有任务完成！所有输出文件已保存至当前代码目录")
        logging.info("=" * 50)

        return {
            "metrics": metrics,
            "threshold_strategy": best_threshold,
            "output_dir": CURRENT_DIR,
            "output_files": output_files,
            "out_prefix": out_prefix,
            "employee_rows": int(len(df_emp)),
            "policy_rows": int(len(policy_df)),
        }
    except SystemExit as exc:
        raise RuntimeError("流程执行被中止，请检查输入文件和运行依赖。") from exc


def main():
    """主流程：数据加载→预处理→特征工程→模型训练→结果输出→可视化"""
    run_pipeline(
        employee_data_path=DATA_PATH,
        policy_data_path=POLICY_PATH,
        output_dir=CURRENT_DIR,
        out_prefix="employee_attrition_analysis",
    )

if __name__ == '__main__':
    main()
